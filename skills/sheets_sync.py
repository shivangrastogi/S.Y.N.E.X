"""Push the local workbook to Google Sheets.

Two modes:
  - Service-account (gspread + a JSON key at data/google_credentials.json)
  - Graceful "creds not set up" message — never raises.

Setup (user does once):
  1. Go to https://console.cloud.google.com → create project
  2. Enable Google Sheets API + Google Drive API
  3. Create a Service Account, download its JSON key
  4. Save the key as: data/google_credentials.json
  5. Either share an existing Sheet with the service account email,
     OR let this skill create a new one (writes the URL into
     data/sheets_sync.json so we remember it).

Voice triggers: "sync to google sheets", "google sheets pe upload karo",
"cloud pe save karo expenses".
"""

from __future__ import annotations

import json
import logging
import os
import sys
from datetime import datetime
from pathlib import Path
from typing import Optional

_HERE = os.path.dirname(os.path.abspath(__file__))
_ROOT = os.path.dirname(_HERE)
if _ROOT not in sys.path:
    sys.path.insert(0, _ROOT)

from core.skill_registry import skill  # noqa: E402

log = logging.getLogger(__name__)

_DATA_DIR = Path(_ROOT) / "data"
_WORKBOOK_PATH = _DATA_DIR / "jarvis_workbook.xlsx"
_CREDS_PATH = _DATA_DIR / "google_credentials.json"
_STATE_PATH = _DATA_DIR / "sheets_sync.json"

_SHEET_TITLE = "AERIS — Life & Expense Workbook"


def _read_state() -> dict:
    if not _STATE_PATH.exists():
        return {}
    try:
        return json.loads(_STATE_PATH.read_text(encoding="utf-8"))
    except Exception:
        return {}


def _write_state(state: dict) -> None:
    _STATE_PATH.parent.mkdir(parents=True, exist_ok=True)
    _STATE_PATH.write_text(json.dumps(state, indent=2), encoding="utf-8")


def _try_clients():
    """Lazy-import gspread + openpyxl. Returns (gspread, openpyxl) or None."""
    try:
        import gspread  # noqa: F401
        import openpyxl  # noqa: F401
        from google.oauth2.service_account import Credentials  # noqa: F401
        return True
    except Exception:
        return False


def _setup_instructions() -> str:
    return (
        "Google Sheets sync ke liye one-time setup chahiye, sir:\n"
        "  1. pip install gspread google-auth openpyxl\n"
        "  2. Google Cloud Console mein Service Account banao\n"
        "  3. JSON key download karke save karo: data/google_credentials.json\n"
        "  4. Phir 'sync to google sheets' bolo — main automatically sheet bana dunga."
    )


def _open_or_create_sheet():
    """Returns (gspread.Spreadsheet, was_created) or raises."""
    import gspread
    from google.oauth2.service_account import Credentials

    scopes = [
        "https://www.googleapis.com/auth/spreadsheets",
        "https://www.googleapis.com/auth/drive",
    ]
    creds = Credentials.from_service_account_file(str(_CREDS_PATH), scopes=scopes)
    gc = gspread.authorize(creds)

    state = _read_state()
    sheet_id = state.get("spreadsheet_id")
    if sheet_id:
        try:
            return gc.open_by_key(sheet_id), False
        except Exception as e:
            log.info("[sheets_sync] cached id failed (%s); creating fresh", e)

    sh = gc.create(_SHEET_TITLE)
    state["spreadsheet_id"] = sh.id
    state["url"] = f"https://docs.google.com/spreadsheets/d/{sh.id}"
    state["created_at"] = datetime.now().isoformat(timespec="seconds")
    _write_state(state)
    return sh, True


def _push_workbook_to_sheet(sh) -> dict:
    """Copy every tab from the local .xlsx to the remote spreadsheet."""
    from openpyxl import load_workbook

    wb = load_workbook(_WORKBOOK_PATH, data_only=False)
    pushed: dict[str, int] = {}

    existing = {ws.title for ws in sh.worksheets()}

    for tab_name in wb.sheetnames:
        ws_local = wb[tab_name]
        rows: list[list] = []
        for r in ws_local.iter_rows(values_only=True):
            row_vals = [
                v.isoformat() if isinstance(v, datetime) else
                (v.isoformat() if hasattr(v, "isoformat") else
                 ("" if v is None else v))
                for v in r
            ]
            rows.append(row_vals)
        if not rows:
            rows = [[""]]

        # Pad rows to a uniform width (Sheets requires it).
        width = max(len(r) for r in rows)
        for r in rows:
            while len(r) < width:
                r.append("")

        if tab_name in existing:
            ws_remote = sh.worksheet(tab_name)
            ws_remote.clear()
        else:
            ws_remote = sh.add_worksheet(
                title=tab_name,
                rows=str(max(len(rows) + 5, 100)),
                cols=str(max(width + 2, 10)),
            )
        ws_remote.update(rows, "A1")
        pushed[tab_name] = len(rows)

    # Drop the gspread-default "Sheet1" if it sneaked in.
    for ws_remote in sh.worksheets():
        if ws_remote.title == "Sheet1" and ws_remote.title not in wb.sheetnames:
            try:
                sh.del_worksheet(ws_remote)
            except Exception:
                pass

    return pushed


@skill(
    name="sync_to_sheets",
    description="Upload the local AERIS workbook to Google Sheets.",
    patterns=[
        "sync to google sheets", "google sheets pe upload karo",
        "cloud pe save karo expenses", "push to sheets",
        "sheets pe sync karo", "google drive pe daal do",
        "expense sheet ko cloud pe daal do",
        "upload workbook to google sheets",
    ],
    required_entities=[],
)
def sync_to_sheets(slots: dict) -> str:
    if not _WORKBOOK_PATH.exists():
        return ("Local workbook abhi bana nahi hai, sir. "
                "Pehle ek expense add karo, phir sync karenge.")

    if not _try_clients():
        return ("gspread / google-auth missing hain. "
                "'pip install gspread google-auth openpyxl' chalao, phir try karo.")

    if not _CREDS_PATH.exists():
        return _setup_instructions()

    try:
        sh, created = _open_or_create_sheet()
    except Exception as e:
        log.warning("[sheets_sync] auth/open failed: %s", e)
        return f"Google Sheets connect nahi ho paya: {e}"

    try:
        pushed = _push_workbook_to_sheet(sh)
    except Exception as e:
        log.warning("[sheets_sync] push failed: %s", e)
        return f"Sync mein dikkat: {e}"

    state = _read_state()
    url = state.get("url", "")
    tab_msg = ", ".join(f"{k}({v})" for k, v in pushed.items())
    intro = "Sheet banayi aur " if created else ""
    return f"{intro}sync ho gaya: {tab_msg}.\n{url}"


@skill(
    name="sheets_status",
    description="Show whether Google Sheets sync is configured and where the synced sheet lives.",
    patterns=[
        "sheets sync status", "google sheets connect hai kya",
        "sheets configured hai kya", "sync status batao",
        "where is my synced sheet",
    ],
    required_entities=[],
)
def sheets_status(slots: dict) -> str:
    has_lib = _try_clients()
    has_creds = _CREDS_PATH.exists()
    state = _read_state()

    lines = [
        f"Library installed: {'yes' if has_lib else 'no — pip install gspread google-auth'}",
        f"Credentials at data/google_credentials.json: "
        f"{'yes' if has_creds else 'no'}",
    ]
    if state.get("url"):
        lines.append(f"Synced sheet: {state['url']}")
    else:
        lines.append("Synced sheet: not yet (run 'sync to google sheets').")
    return "\n".join(lines)


if __name__ == "__main__":
    logging.basicConfig(level=logging.INFO, format="%(message)s")
    print(sheets_status({}))
