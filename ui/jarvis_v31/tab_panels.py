"""JARVIS v3.1 — right-panel tab pages.

One glass panel per dock tab. Each is 390px wide and uses the same
translucent-dark design language as GlassChatPanel. Backend wiring
is not implemented — these are complete UI shells.

Tabs: home · chat · auto · brain · memory · system · settings
"""
from __future__ import annotations

import math
import platform
import sys
import time

from PyQt5.QtCore import QPointF, QRectF, Qt, QTimer, pyqtSignal as pyqtSignal
from PyQt5.QtGui import (
    QBrush, QColor, QLinearGradient, QPainter, QPen,
)
from PyQt5.QtWidgets import (
    QComboBox, QFrame, QGridLayout, QHBoxLayout, QLabel, QLineEdit,
    QPushButton, QScrollArea, QStackedWidget, QVBoxLayout, QWidget,
)

from .animation_bus import get_bus
from .tokens import J, inter, mono, rgba


# ─── Shared primitives ────────────────────────────────────────── #

class _GlassPanel(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setFixedWidth(390)

    def paintEvent(self, _):
        p = QPainter(self)
        bg = QColor(J.PANEL); bg.setAlphaF(0.75)
        p.fillRect(self.rect(), bg)
        p.setPen(QPen(rgba(J.BORDER, 0.6), 1))
        p.drawLine(0, 0, 0, self.height())
        p.setPen(QPen(rgba(J.CYAN, 0.06), 1))
        p.drawLine(1, 0, 1, self.height())


class _PanelHeader(QFrame):
    def __init__(self, title: str, sub: str, color: QColor, parent=None):
        super().__init__(parent)
        self.setFixedHeight(56)
        self._color = color
        lay = QHBoxLayout(self)
        lay.setContentsMargins(20, 0, 16, 0); lay.setSpacing(8)
        left = QVBoxLayout(); left.setSpacing(2)
        t = QLabel(title); t.setFont(inter(11, 700))
        t.setStyleSheet(f"color:{J.TEXT_PRI.name()};letter-spacing:2px;")
        s = QLabel(sub); s.setFont(mono(10, 400))
        s.setStyleSheet(f"color:{J.TEXT_MUT.name()};")
        left.addWidget(t); left.addWidget(s)
        lay.addLayout(left); lay.addStretch(1)
        lay.addWidget(_BlinkDot(color))

    def paintEvent(self, _):
        p = QPainter(self)
        bg = QColor(J.BG); bg.setAlphaF(0.4)
        p.fillRect(self.rect(), bg)
        p.setPen(QPen(rgba(J.BORDER, 0.7), 1))
        p.drawLine(0, self.height()-1, self.width(), self.height()-1)


class _BlinkDot(QWidget):
    def __init__(self, color: QColor, parent=None):
        super().__init__(parent)
        self._color = color
        self.setFixedSize(8, 8)
        self._bus = get_bus()
        self._bus.tick_slow.connect(self.update)

    def paintEvent(self, _):
        ph = self._bus.now_ms / 1000.0 * 2 * math.pi / 1.8
        a = 0.4+0.6*(0.5+0.5*math.sin(ph))
        p = QPainter(self); p.setRenderHint(QPainter.Antialiasing, True)
        p.setPen(Qt.NoPen)
        p.setBrush(rgba(self._color, a*0.35)); p.drawEllipse(self.rect())
        p.setBrush(rgba(self._color, a)); p.drawEllipse(self.rect().adjusted(2,2,-2,-2))


def _scrollable(content: QWidget) -> QScrollArea:
    sa = QScrollArea(); sa.setWidgetResizable(True); sa.setFrameShape(QFrame.NoFrame)
    sa.setStyleSheet(
        "QScrollArea{background:transparent;border:none;}"
        "QScrollBar:vertical{background:transparent;width:4px;}"
        "QScrollBar::handle:vertical{background:rgba(0,212,255,0.18);border-radius:2px;}"
        "QScrollBar::add-line,QScrollBar::sub-line{height:0;}"
    )
    content.setStyleSheet("background:transparent;")
    sa.setWidget(content); return sa


def _sec(text: str) -> QLabel:
    l = QLabel(text); l.setFont(mono(9, 700))
    l.setStyleSheet(f"color:{J.TEXT_MUT.name()};letter-spacing:1.5px;")
    return l


def _divider() -> QWidget:
    w = QWidget(); w.setFixedHeight(1)
    w.setStyleSheet(f"background:rgba({J.BORDER.red()},{J.BORDER.green()},{J.BORDER.blue()},0.30);")
    return w


class _Card(QWidget):
    def __init__(self, color: QColor = None, parent=None):
        super().__init__(parent)
        self._color = color or J.BORDER

    def paintEvent(self, _):
        p = QPainter(self); p.setRenderHint(QPainter.Antialiasing, True)
        bg = QColor(J.BG_ELE); bg.setAlphaF(0.60)
        p.setPen(QPen(rgba(self._color, 0.22), 1))
        p.setBrush(bg)
        p.drawRoundedRect(QRectF(0.5,0.5,self.width()-1,self.height()-1), 10, 10)


class _StatCard(QWidget):
    def __init__(self, value: str, label: str, color: QColor, parent=None):
        super().__init__(parent)
        self._color = color
        lay = QVBoxLayout(self); lay.setContentsMargins(12,10,12,10); lay.setSpacing(3)
        self._value_lbl = QLabel(value); self._value_lbl.setFont(inter(22, 700))
        self._value_lbl.setStyleSheet(f"color:{color.name()};")
        self._value_lbl.setAlignment(Qt.AlignCenter)
        l = QLabel(label); l.setFont(mono(9, 400))
        l.setStyleSheet(f"color:{J.TEXT_MUT.name()};letter-spacing:0.4px;")
        l.setAlignment(Qt.AlignCenter); l.setWordWrap(True)
        lay.addWidget(self._value_lbl); lay.addWidget(l)

    def set_value(self, v: str) -> None:
        """Update the big-number label live (used by HomePanel)."""
        self._value_lbl.setText(str(v))

    def paintEvent(self, _):
        p = QPainter(self); p.setRenderHint(QPainter.Antialiasing, True)
        bg = QColor(J.BG_ELE); bg.setAlphaF(0.55)
        p.setPen(QPen(rgba(self._color, 0.20), 1))
        p.setBrush(bg)
        p.drawRoundedRect(QRectF(0.5,0.5,self.width()-1,self.height()-1), 10, 10)


class _StatusRow(QWidget):
    def __init__(self, label: str, status: str, ok: bool = True, parent=None):
        super().__init__(parent)
        lay = QHBoxLayout(self); lay.setContentsMargins(0, 4, 0, 4); lay.setSpacing(8)
        col = J.GREEN if ok else J.RED
        dot = QWidget(); dot.setFixedSize(7, 7)
        dot.setStyleSheet(f"background:{col.name()};border-radius:3px;")
        lay.addWidget(dot, 0, Qt.AlignVCenter)
        lbl = QLabel(label); lbl.setFont(mono(10, 400))
        lbl.setStyleSheet(f"color:{J.TEXT_SEC.name()};")
        lay.addWidget(lbl, 1)
        st = QLabel(status); st.setFont(mono(9, 700))
        st.setStyleSheet(f"color:{col.name()};")
        lay.addWidget(st)


class _Toggle(QWidget):
    toggled = pyqtSignal(bool)

    def __init__(self, on: bool = False, parent=None):
        super().__init__(parent)
        self._on = on
        self.setFixedSize(42, 24)
        self.setCursor(Qt.PointingHandCursor)
        # No timer — toggle is static between clicks; repaint on press only.

    def mousePressEvent(self, e):
        if e.button() == Qt.LeftButton:
            self._on = not self._on
            self.toggled.emit(self._on)
            self.update()

    def paintEvent(self, _):
        p = QPainter(self); p.setRenderHint(QPainter.Antialiasing, True)
        p.setPen(Qt.NoPen)
        track = J.CYAN if self._on else rgba(J.BORDER, 1.0)
        p.setBrush(rgba(track, 0.35 if self._on else 0.20))
        p.drawRoundedRect(QRectF(0, 5, 42, 14), 7, 7)
        cx = 32.0 if self._on else 10.0
        p.setBrush(J.CYAN if self._on else J.TEXT_MUT)
        p.drawEllipse(QPointF(cx, 12), 8, 8)


class _ToggleRow(QWidget):
    def __init__(self, label: str, sub: str, on: bool, parent=None):
        super().__init__(parent)
        lay = QHBoxLayout(self); lay.setContentsMargins(0, 4, 0, 4); lay.setSpacing(10)
        left = QVBoxLayout(); left.setSpacing(1)
        t = QLabel(label); t.setFont(inter(12, 500))
        t.setStyleSheet(f"color:{J.TEXT_PRI.name()};")
        left.addWidget(t)
        if sub:
            s = QLabel(sub); s.setFont(mono(9, 400))
            s.setStyleSheet(f"color:{J.TEXT_MUT.name()};")
            left.addWidget(s)
        lay.addLayout(left); lay.addStretch(1)
        lay.addWidget(_Toggle(on))


# ─── Home panel ───────────────────────────────────────────────── #

class HomePanel(_GlassPanel):
    """Live dashboard — replaces the original static shell.

    Refreshes on tick_slow (throttled to 1 Hz internally). Pulls data
    from ResourceMonitor, PowerMonitor, WorkspaceMonitor, automation
    engine, and log_setup's event ring buffer — everything we already
    expose elsewhere, just composed.
    """

    def __init__(self, parent=None):
        super().__init__(parent)
        self._bus = get_bus()
        self._bus.tick_slow.connect(self._maybe_refresh)
        self._last_refresh = 0.0
        self._wx_cache_text = ""
        self._wx_cache_ts = 0.0

        root = QVBoxLayout(self); root.setContentsMargins(0,0,0,0); root.setSpacing(0)
        root.addWidget(_PanelHeader("JARVIS HOME", "Dashboard · welcome back", J.CYAN))

        inner = QWidget()
        lay = QVBoxLayout(inner); lay.setContentsMargins(16, 16, 16, 18); lay.setSpacing(14)

        # ── Big clock + date ──
        self._clock = QLabel("--:--")
        self._clock.setFont(mono(36, 700))
        self._clock.setStyleSheet(f"color:{J.CYAN.name()};letter-spacing:2px;")
        self._clock.setAlignment(Qt.AlignCenter)
        lay.addWidget(self._clock)
        self._date = QLabel("—")
        self._date.setFont(mono(10, 400))
        self._date.setStyleSheet(f"color:{J.TEXT_MUT.name()};letter-spacing:1px;")
        self._date.setAlignment(Qt.AlignCenter)
        lay.addWidget(self._date)

        # ── Workspace + battery row ──
        self._workspace = QLabel("workspace: —")
        self._workspace.setFont(mono(9, 700))
        self._workspace.setStyleSheet(f"color:{J.AMBER.name()};letter-spacing:1px;")
        self._workspace.setAlignment(Qt.AlignCenter)
        lay.addWidget(self._workspace)

        lay.addWidget(_divider())

        # ── Weather card ──
        lay.addWidget(_sec("WEATHER"))
        wx_card = _Card(J.CYAN)
        wcl = QVBoxLayout(wx_card); wcl.setContentsMargins(14, 10, 14, 10); wcl.setSpacing(2)
        self._wx_label = QLabel("Loading...")
        self._wx_label.setFont(mono(10, 400))
        self._wx_label.setStyleSheet(f"color:{J.TEXT_SEC.name()};")
        self._wx_label.setWordWrap(True)
        wcl.addWidget(self._wx_label)
        lay.addWidget(wx_card)

        # ── Stats grid (real numbers) ──
        lay.addWidget(_sec("LIVE STATS"))
        self._stat_grid = QGridLayout(); self._stat_grid.setSpacing(8)
        self._stats: dict[str, _StatCard] = {}
        for i, (key, label, color) in enumerate([
            ("ram",     "RAM\nMB",       J.CYAN),
            ("cpu",     "CPU\n%",        J.GREEN),
            ("battery", "Battery\n%",    J.AMBER),
            ("skills",  "Skills\nLoaded", J.PURPLE),
        ]):
            card = _StatCard("—", label, color)
            self._stats[key] = card
            self._stat_grid.addWidget(card, i // 2, i % 2)
        lay.addLayout(self._stat_grid)

        lay.addWidget(_divider())

        # ── Routines glance ──
        lay.addWidget(_sec("ENABLED ROUTINES"))
        self._routines_label = QLabel("—")
        self._routines_label.setFont(mono(9, 400))
        self._routines_label.setStyleSheet(f"color:{J.TEXT_SEC.name()};")
        self._routines_label.setWordWrap(True)
        lay.addWidget(self._routines_label)

        lay.addWidget(_divider())

        # ── Recent events feed (last 6) ──
        lay.addWidget(_sec("RECENT EVENTS"))
        self._events_label = QLabel("(none yet)")
        self._events_label.setFont(mono(9, 400))
        self._events_label.setStyleSheet(f"color:{J.TEXT_MUT.name()};")
        self._events_label.setWordWrap(True)
        lay.addWidget(self._events_label)

        lay.addStretch(1)
        root.addWidget(_scrollable(inner), 1)

        # First paint — soon after construction so the widget isn't blank.
        QTimer.singleShot(120, self._refresh)

    # ── tick ───────────────────────────────────────────────────────

    def _maybe_refresh(self):
        now = time.monotonic()
        if now - self._last_refresh < 1.0:
            return
        try:
            self._refresh()
        except Exception:
            pass

    def _refresh(self):
        self._last_refresh = time.monotonic()
        now = time.localtime()
        self._clock.setText(time.strftime("%H:%M", now))
        self._date.setText(time.strftime("%A · %d %b %Y", now))

        # Workspace + battery banner
        ws = "—"
        try:
            from core.workspace_profile import get_monitor as _gw
            ws = _gw().snapshot().profile.lower()
        except Exception:
            pass
        bat_str = ""
        try:
            from core.power_state import get_monitor as _gp
            snap = _gp().snapshot()
            src = "battery" if snap.on_battery else "AC"
            bat_str = f"  ·  {snap.percent}% {src}"
        except Exception:
            pass
        self._workspace.setText(f"workspace: {ws}{bat_str}")

        # Live stats
        try:
            from core.resource_monitor import get_monitor as _gm
            samples = _gm().snapshot()
            if samples:
                last = samples[-1]
                self._stats["ram"].set_value(f"{last.rss_bytes / (1024*1024):.0f}")
                self._stats["cpu"].set_value(f"{last.cpu_percent:.0f}")
        except Exception:
            pass
        try:
            from core.power_state import get_monitor as _gp
            self._stats["battery"].set_value(str(_gp().snapshot().percent))
        except Exception:
            pass
        try:
            from core.skill_registry import REGISTRY
            self._stats["skills"].set_value(str(len(REGISTRY)))
        except Exception:
            pass

        # Weather — cached 10 min, fetched lazily on first paint.
        if time.monotonic() - self._wx_cache_ts > 600 or not self._wx_cache_text:
            # Trigger async fetch via QTimer so we never block the GUI thread.
            QTimer.singleShot(0, self._fetch_weather)
        if self._wx_cache_text:
            self._wx_label.setText(self._wx_cache_text)

        # Routines glance
        try:
            from core.automation import get_engine
            routines = [r for r in get_engine().list_routines() if r.enabled]
            if not routines:
                self._routines_label.setText("(none enabled)")
            else:
                lines = [f"  · {r.name}  —  {r.trigger.summary()}"
                         for r in routines[:4]]
                if len(routines) > 4:
                    lines.append(f"  + {len(routines) - 4} more")
                self._routines_label.setText("\n".join(lines))
        except Exception:
            self._routines_label.setText("(engine offline)")

        # Recent events
        try:
            from core.log_setup import snapshot_events
            from datetime import datetime as _dt
            evs = snapshot_events()[-6:]
            if not evs:
                self._events_label.setText("(none yet)")
            else:
                lines = []
                for ev in reversed(evs):
                    ts = _dt.fromtimestamp(ev.get("ts", 0)).strftime("%H:%M:%S")
                    name = ev.get("event", "?")
                    extras = " ".join(f"{k}={v}" for k, v in ev.items()
                                      if k not in ("ts", "event"))
                    suffix = f"  ·  {extras}" if extras else ""
                    lines.append(f"  {ts}  {name}{suffix}"[:80])
                self._events_label.setText("\n".join(lines))
        except Exception:
            self._events_label.setText("(no event bus)")

    def _fetch_weather(self):
        """Pulled out so the GUI thread doesn't block on the HTTP call.
        On Windows this still runs on the GUI thread but in a separate
        slot, so input events get a chance between the timer fire and
        the network call.
        """
        try:
            from skills.info_skills import weather_local
            txt = weather_local({})
            self._wx_cache_text = txt
            self._wx_cache_ts = time.monotonic()
            self._wx_label.setText(txt)
        except Exception as e:
            self._wx_label.setText(f"(weather unavailable: {e})")


# ─── Brain panel ──────────────────────────────────────────────── #

class BrainPanel(_GlassPanel):
    def __init__(self, parent=None):
        super().__init__(parent)
        root = QVBoxLayout(self); root.setContentsMargins(0,0,0,0); root.setSpacing(0)
        root.addWidget(_PanelHeader("NEURAL BRAIN", "Intent engine · model info", J.MAGENTA))

        inner = QWidget()
        lay = QVBoxLayout(inner); lay.setContentsMargins(16,18,16,18); lay.setSpacing(16)

        lay.addWidget(_sec("MODEL STACK"))
        card = _Card(J.MAGENTA)
        cl = QVBoxLayout(card); cl.setContentsMargins(14,12,14,12); cl.setSpacing(7)
        for k, v in (
            ("Encoder",    "paraphrase-multilingual-MiniLM-L12-v2"),
            ("Classifier", "NearestNeighbors · cosine · k=5"),
            ("Voting",     "exp-weighted · temperature=10"),
            ("Cache",      "data/models/intent_index.pkl"),
            ("Intents",    "21 classes · ~14 Hinglish patterns each"),
        ):
            rw = QHBoxLayout(); rw.setSpacing(8)
            kl = QLabel(k); kl.setFont(mono(9, 700))
            kl.setStyleSheet(f"color:{J.MAGENTA.name()};"); kl.setFixedWidth(72)
            vl = QLabel(v); vl.setFont(mono(9, 400))
            vl.setStyleSheet(f"color:{J.TEXT_SEC.name()};"); vl.setWordWrap(True)
            rw.addWidget(kl); rw.addWidget(vl, 1)
            cl.addLayout(rw)
        lay.addWidget(card)

        lay.addWidget(_divider())
        lay.addWidget(_sec("ACTIVE INTENTS  ·  21"))
        for intent in ("greet", "farewell", "weather_check", "open_app",
                       "play_music", "system_stats", "time_query", "chit_chat",
                       "set_timer", "take_note", "recall_fact", "+ 10 more"):
            rw = QHBoxLayout(); rw.setContentsMargins(0, 2, 0, 2)
            n = QLabel(intent); n.setFont(mono(10, 400))
            n.setStyleSheet(f"color:{J.TEXT_SEC.name()};")
            rw.addWidget(n); rw.addStretch(1)
            col = J.TEXT_MUT if intent.startswith("+") else J.GREEN
            b = QLabel("—" if intent.startswith("+") else "ACTIVE"); b.setFont(mono(8, 700))
            b.setStyleSheet(f"color:{col.name()};")
            rw.addWidget(b)
            lay.addLayout(rw)

        lay.addStretch(1)
        root.addWidget(_scrollable(inner), 1)


# ─── Memory panel ─────────────────────────────────────────────── #

class MemoryPanel(_GlassPanel):
    def __init__(self, parent=None):
        super().__init__(parent)
        root = QVBoxLayout(self); root.setContentsMargins(0,0,0,0); root.setSpacing(0)
        root.addWidget(_PanelHeader("USER MEMORY", "Persistent facts · learned data", J.PURPLE))

        inner = QWidget()
        lay = QVBoxLayout(inner); lay.setContentsMargins(16,18,16,18); lay.setSpacing(16)

        grid = QGridLayout(); grid.setSpacing(8)
        for i, (v, l, c) in enumerate([
            ("0", "Total Facts",  J.PURPLE),
            ("0", "Categories",   J.CYAN),
        ]):
            grid.addWidget(_StatCard(v, l, c), 0, i)
        lay.addLayout(grid)

        lay.addWidget(_divider())
        lay.addWidget(_sec("STORED FACTS"))

        card = _Card(J.PURPLE)
        ecl = QVBoxLayout(card); ecl.setContentsMargins(14, 24, 14, 24); ecl.setSpacing(10)
        icon = QLabel("◈"); icon.setFont(inter(30, 700))
        icon.setStyleSheet(f"color:{J.PURPLE.name()};")
        icon.setAlignment(Qt.AlignCenter)
        ecl.addWidget(icon)
        msg = QLabel("No facts stored yet.\nTell me something —\ne.g. \"My name is Shivang\"")
        msg.setFont(mono(10, 400)); msg.setAlignment(Qt.AlignCenter)
        msg.setStyleSheet(f"color:{J.TEXT_MUT.name()};"); msg.setWordWrap(True)
        ecl.addWidget(msg)
        lay.addWidget(card)

        lay.addWidget(_divider())
        lay.addWidget(_sec("HOW IT WORKS"))
        for tip in (
            "Say facts naturally in conversation",
            "\"My name is…\" · \"I prefer…\" · \"I work at…\"",
            "Jarvis remembers across sessions",
            "Ask \"what do you know about me?\"",
        ):
            tl = QLabel(f"  {tip}"); tl.setFont(mono(9, 400))
            tl.setStyleSheet(f"color:{J.TEXT_MUT.name()};")
            lay.addWidget(tl)

        lay.addStretch(1)
        root.addWidget(_scrollable(inner), 1)


# ─── System panel ─────────────────────────────────────────────── #

class _LiveMetrics(QWidget):
    """Self-refreshing metrics block — RSS history, caches, breakers, power.

    Pulls fresh state from the in-process registries every ~1s (tied to
    tick_slow, ~15 FPS, throttled internally). All sources are
    *optional* — missing modules degrade gracefully to a "–" reading.
    """

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setMinimumHeight(140)
        self._bus = get_bus()
        self._bus.tick_slow.connect(self._tick)
        self._last_refresh = 0.0

        lay = QVBoxLayout(self); lay.setContentsMargins(0,0,0,0); lay.setSpacing(7)
        # RSS row
        self._rss = QLabel("RSS — MB"); self._rss.setFont(mono(10, 700))
        self._rss.setStyleSheet(f"color:{J.CYAN.name()};")
        lay.addWidget(self._rss)
        self._spark = _RssSparkline()
        lay.addWidget(self._spark)
        # CPU + threads + handles row
        self._stat = QLabel("CPU — % · threads — · handles —")
        self._stat.setFont(mono(9, 400))
        self._stat.setStyleSheet(f"color:{J.TEXT_SEC.name()};")
        lay.addWidget(self._stat)
        # Pressure pill
        self._pressure = QLabel("pressure: OK")
        self._pressure.setFont(mono(9, 700))
        self._pressure.setStyleSheet(f"color:{J.GREEN.name()};")
        lay.addWidget(self._pressure)
        # Power row
        self._power = QLabel("power: —")
        self._power.setFont(mono(9, 400))
        self._power.setStyleSheet(f"color:{J.TEXT_SEC.name()};")
        lay.addWidget(self._power)
        # Cache row
        self._caches = QLabel("caches: —")
        self._caches.setFont(mono(9, 400))
        self._caches.setStyleSheet(f"color:{J.TEXT_SEC.name()};")
        lay.addWidget(self._caches)
        # Breakers row
        self._breakers = QLabel("skill breakers: —")
        self._breakers.setFont(mono(9, 400))
        self._breakers.setStyleSheet(f"color:{J.TEXT_SEC.name()};")
        lay.addWidget(self._breakers)

    def _tick(self):
        # Throttle to 1 Hz — tick_slow fires ~15 Hz which is overkill.
        now = time.monotonic()
        if now - self._last_refresh < 1.0:
            return
        self._last_refresh = now
        try:
            self._refresh()
        except Exception:
            pass

    def _refresh(self):
        # ── ResourceMonitor ──
        rss_mb = 0.0; cpu = 0.0; threads = 0; handles = 0; pressure = "OK"
        samples_mb: list[float] = []
        try:
            from core.resource_monitor import get_monitor as _gm
            m = _gm()
            samples = m.snapshot()
            if samples:
                last = samples[-1]
                rss_mb = last.rss_bytes / (1024 * 1024)
                cpu = last.cpu_percent
                threads = last.num_threads
                handles = last.num_handles
            samples_mb = [s.rss_bytes / (1024 * 1024) for s in samples[-60:]]
            pressure = m.level_name()
        except Exception:
            pass
        self._rss.setText(f"RSS {rss_mb:.0f} MB")
        self._stat.setText(f"CPU {cpu:.1f}% · threads {threads} · handles {handles}")
        self._spark.set_data(samples_mb)
        pcol = J.GREEN if pressure == "OK" else (J.AMBER if pressure == "WARNING" else J.RED)
        self._pressure.setText(f"pressure: {pressure}")
        self._pressure.setStyleSheet(f"color:{pcol.name()};")

        # ── Power ──
        try:
            from core.power_state import get_monitor as _gpm
            snap = _gpm().snapshot()
            bat = "battery" if snap.on_battery else "AC"
            self._power.setText(
                f"power: {bat} {snap.percent}% · idle {snap.idle_s}s"
                + (" · IDLE" if snap.is_idle else "")
            )
        except Exception:
            self._power.setText("power: —")

        # ── Caches ──
        try:
            from core.cache_registry import cache_stats
            cs = cache_stats()
            if cs:
                hits = sum(c["hits"] for c in cs)
                miss = sum(c["misses"] for c in cs)
                rate = (hits / (hits + miss) * 100) if (hits + miss) else 0.0
                self._caches.setText(
                    f"caches: {len(cs)} · {hits + miss} calls · {rate:.0f}% hit rate"
                )
            else:
                self._caches.setText("caches: 0")
        except Exception:
            self._caches.setText("caches: —")

        # ── Skill breakers ──
        try:
            from core.skill_breaker import stats as _bs
            bs = _bs()
            if bs:
                opens = sum(1 for b in bs if b["state"] == "OPEN")
                self._breakers.setText(
                    f"skill breakers: {len(bs)} tracked · {opens} OPEN"
                )
                if opens > 0:
                    self._breakers.setStyleSheet(f"color:{J.RED.name()};")
                else:
                    self._breakers.setStyleSheet(f"color:{J.TEXT_SEC.name()};")
            else:
                self._breakers.setText("skill breakers: 0")
        except Exception:
            self._breakers.setText("skill breakers: —")


class _RssSparkline(QWidget):
    """Tiny inline graph of the last ~60 RSS samples in MB."""

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setMinimumHeight(28)
        self.setMaximumHeight(28)
        self._data: list[float] = []

    def set_data(self, data: list[float]) -> None:
        self._data = list(data)
        self.update()

    def paintEvent(self, _):
        p = QPainter(self); p.setRenderHint(QPainter.Antialiasing, True)
        w, h = self.width(), self.height()
        # background line
        p.setPen(QPen(rgba(J.CYAN, 0.18), 1))
        p.drawLine(0, h - 1, w, h - 1)
        if not self._data:
            return
        lo = min(self._data); hi = max(self._data)
        rng = (hi - lo) if hi > lo else 1.0
        n = len(self._data)
        step = w / max(1, n - 1) if n > 1 else w
        pts = []
        for i, v in enumerate(self._data):
            x = i * step
            y = (h - 4) - ((v - lo) / rng) * (h - 8)
            pts.append((x, y))
        # filled area
        from PyQt5.QtGui import QPainterPath
        path = QPainterPath()
        path.moveTo(0, h)
        for x, y in pts: path.lineTo(x, y)
        path.lineTo(w, h); path.lineTo(0, h)
        grad = QLinearGradient(0, 0, 0, h)
        grad.setColorAt(0.0, rgba(J.CYAN, 0.45))
        grad.setColorAt(1.0, rgba(J.CYAN, 0.05))
        p.setPen(Qt.NoPen); p.setBrush(grad); p.drawPath(path)
        # line on top
        p.setPen(QPen(J.CYAN, 1.4))
        for i in range(len(pts) - 1):
            x0, y0 = pts[i]; x1, y1 = pts[i + 1]
            p.drawLine(QPointF(x0, y0), QPointF(x1, y1))


class SystemPanel(_GlassPanel):
    def __init__(self, parent=None):
        super().__init__(parent)
        root = QVBoxLayout(self); root.setContentsMargins(0,0,0,0); root.setSpacing(0)
        root.addWidget(_PanelHeader("SYSTEM STATUS", "Runtime · module health", J.GREEN))

        inner = QWidget()
        lay = QVBoxLayout(inner); lay.setContentsMargins(16,18,16,18); lay.setSpacing(16)

        # ── Live metrics (top) — pulled from ResourceMonitor + friends ──
        lay.addWidget(_sec("LIVE METRICS"))
        live_card = _Card(J.CYAN)
        lcl = QVBoxLayout(live_card); lcl.setContentsMargins(14,12,14,12); lcl.setSpacing(6)
        lcl.addWidget(_LiveMetrics())
        lay.addWidget(live_card)
        lay.addWidget(_divider())

        lay.addWidget(_sec("RUNTIME"))
        card = _Card(J.GREEN)
        cl = QVBoxLayout(card); cl.setContentsMargins(14,12,14,12); cl.setSpacing(7)
        try:
            pv = f"{sys.version_info.major}.{sys.version_info.minor}.{sys.version_info.micro}"
        except Exception:
            pv = "?"
        for k, v in (
            ("OS",      platform.system() + " " + platform.release()),
            ("Python",  pv),
            ("Arch",    platform.machine()),
            ("Build",   "JARVIS v3.1 · 2025"),
        ):
            rw = QHBoxLayout(); rw.setSpacing(8)
            kl = QLabel(k); kl.setFont(mono(9, 700))
            kl.setStyleSheet(f"color:{J.GREEN.name()};"); kl.setFixedWidth(52)
            vl = QLabel(v); vl.setFont(mono(9, 400))
            vl.setStyleSheet(f"color:{J.TEXT_SEC.name()};")
            rw.addWidget(kl); rw.addWidget(vl, 1)
            cl.addLayout(rw)
        lay.addWidget(card)

        lay.addWidget(_divider())
        lay.addWidget(_sec("MODULE STATUS"))
        for label, status, ok in (
            ("NLU Brain",     "online",   True),
            ("Edge TTS",      "online",   True),
            ("Google STT",    "online",   True),
            ("Vosk Fallback", "standby",  True),
            ("Ollama LLM",    "offline",  False),
            ("Memory DB",     "online",   True),
            ("Feedback Log",  "online",   True),
            ("Entity NER",    "online",   True),
        ):
            lay.addWidget(_StatusRow(label, status, ok))

        lay.addStretch(1)
        root.addWidget(_scrollable(inner), 1)


# ─── Settings panel ───────────────────────────────────────────── #

class SettingsPanel(_GlassPanel):
    def __init__(self, parent=None):
        super().__init__(parent)
        root = QVBoxLayout(self); root.setContentsMargins(0,0,0,0); root.setSpacing(0)
        root.addWidget(_PanelHeader("SETTINGS", "Preferences · configuration", J.AMBER))

        inner = QWidget()
        lay = QVBoxLayout(inner); lay.setContentsMargins(16,18,16,18); lay.setSpacing(14)

        # ── Identity ──
        from core import settings as _settings
        lay.addWidget(_sec("IDENTITY"))

        name_row = QWidget()
        nl = QHBoxLayout(name_row); nl.setContentsMargins(0, 0, 0, 0); nl.setSpacing(8)
        n_lbl = QLabel("Name"); n_lbl.setFont(mono(10, 400))
        n_lbl.setStyleSheet(f"color:{J.TEXT_MUT.name()};"); n_lbl.setFixedWidth(70)
        self._name_edit = QLineEdit(_settings.assistant_name())
        self._name_edit.setFont(inter(11, 500))
        self._name_edit.setStyleSheet(
            f"QLineEdit{{background:rgba(0,0,0,0.35);color:{J.TEXT_PRI.name()};"
            f"border:1px solid {J.BORDER.name()};border-radius:6px;padding:6px 9px;}}"
            f"QLineEdit:focus{{border-color:{J.CYAN.name()};}}"
        )
        self._name_edit.editingFinished.connect(self._save_name)
        nl.addWidget(n_lbl); nl.addWidget(self._name_edit, 1)
        lay.addWidget(name_row)

        # ── Browser ──
        lay.addWidget(_divider())
        lay.addWidget(_sec("DEFAULT BROWSER"))

        br_row = QWidget()
        bl = QHBoxLayout(br_row); bl.setContentsMargins(0, 0, 0, 0); bl.setSpacing(8)
        b_lbl = QLabel("Browser"); b_lbl.setFont(mono(10, 400))
        b_lbl.setStyleSheet(f"color:{J.TEXT_MUT.name()};"); b_lbl.setFixedWidth(70)
        self._browser_combo = QComboBox()
        self._browser_combo.setFont(inter(11, 500))
        self._browser_combo.setStyleSheet(
            f"QComboBox{{background:rgba(0,0,0,0.35);color:{J.TEXT_PRI.name()};"
            f"border:1px solid {J.BORDER.name()};border-radius:6px;padding:5px 8px;}}"
            f"QComboBox:focus{{border-color:{J.CYAN.name()};}}"
        )
        try:
            from core.browser_launcher import list_known_browsers
            choices = ["system"] + list_known_browsers()
        except Exception:
            choices = ["system", "chrome", "edge", "brave", "firefox", "opera"]
        self._browser_combo.addItems(choices)
        cur = _settings.get("default_browser", "system")
        if cur in choices:
            self._browser_combo.setCurrentText(cur)
        self._browser_combo.currentTextChanged.connect(self._save_browser)
        bl.addWidget(b_lbl); bl.addWidget(self._browser_combo, 1)
        lay.addWidget(br_row)

        hint = QLabel("Used for 'search' / 'open URL' when you don't name a browser.")
        hint.setFont(mono(9, 400)); hint.setWordWrap(True)
        hint.setStyleSheet(f"color:{J.TEXT_MUT.name()};")
        lay.addWidget(hint)

        # ── Voice ──
        lay.addWidget(_divider())
        lay.addWidget(_sec("VOICE"))
        for lbl, sub, on in (
            ("Speech-to-Text",  "Google STT · en-IN + Vosk fallback",  True),
            ("Text-to-Speech",  "Edge-TTS Neerja · pyttsx3 fallback",  True),
            ("Wake Word",       "Say the assistant name to activate",   False),
            ("Auto Sleep",      "Pauses mic after 30 s silence",        True),
        ):
            lay.addWidget(_ToggleRow(lbl, sub, on))

        lay.addWidget(_divider())
        lay.addWidget(_sec("INTELLIGENCE"))
        for lbl, sub, on in (
            ("Hinglish Mode",    "Mixed Hindi + English input",          True),
            ("LLM Fallback",     "Ollama chit-chat for unknown intents", False),
            ("Feedback Learn",   "Improve thresholds from corrections",  True),
            ("Disambiguate",     "Ask when top intents are close",       True),
        ):
            lay.addWidget(_ToggleRow(lbl, sub, on))

        lay.addWidget(_divider())
        lay.addWidget(_sec("INTERFACE"))
        for lbl, sub, on in (
            ("Debug Logs",    "Show verbose logs bar at bottom",  True),
            ("Animations",    "Particle field + reactor rings",   True),
            ("Sound FX",      "UI sounds (coming soon)",          False),
        ):
            lay.addWidget(_ToggleRow(lbl, sub, on))

        lay.addWidget(_divider())
        lay.addWidget(_sec("ABOUT"))
        about_name = QLabel(f"{_settings.assistant_name()} v3.4  ·  build 2026")
        about_name.setFont(mono(9, 400))
        about_name.setStyleSheet(f"color:{J.TEXT_MUT.name()};")
        lay.addWidget(about_name)
        self._about_name_label = about_name
        for line in (
            "Powered by sentence-transformers + Qt5",
            "Hinglish-native · fully local",
        ):
            l = QLabel(line); l.setFont(mono(9, 400))
            l.setStyleSheet(f"color:{J.TEXT_MUT.name()};")
            lay.addWidget(l)

        lay.addStretch(1)
        root.addWidget(_scrollable(inner), 1)

    def _save_name(self) -> None:
        from core import settings as _settings
        new = (self._name_edit.text() or "").strip()
        if not new:
            self._name_edit.setText(_settings.assistant_name())
            return
        _settings.set_("assistant_name", new)
        _settings.set_("wake_word", new.lower())
        if hasattr(self, "_about_name_label"):
            self._about_name_label.setText(f"{new} v3.4  ·  build 2026")

    def _save_browser(self, value: str) -> None:
        from core import settings as _settings
        _settings.set_("default_browser", value)


# ─── Vision panel ─────────────────────────────────────────────── #

class VisionPanel(_GlassPanel):
    """Live status of the gesture engine + object detection toggles."""

    def __init__(self, parent=None):
        super().__init__(parent)
        root = QVBoxLayout(self); root.setContentsMargins(0, 0, 0, 0); root.setSpacing(0)
        root.addWidget(_PanelHeader("VISION CORTEX", "Gestures · objects · webcam", J.GREEN))

        inner = QWidget()
        lay = QVBoxLayout(inner); lay.setContentsMargins(16, 18, 16, 18); lay.setSpacing(16)

        # Status card
        lay.addWidget(_sec("LIVE STATUS"))
        self._status_card = _Card(J.GREEN)
        scl = QVBoxLayout(self._status_card)
        scl.setContentsMargins(14, 12, 14, 12); scl.setSpacing(7)
        self._cam_row = _StatusRow("Camera",     "OFFLINE", ok=False)
        self._gest_row = _StatusRow("Gestures",   "OFF",     ok=False)
        self._obj_row = _StatusRow("Object detector", "READY", ok=True)
        scl.addWidget(self._cam_row)
        scl.addWidget(self._gest_row)
        scl.addWidget(self._obj_row)
        lay.addWidget(self._status_card)

        # Gesture cheat-sheet
        lay.addWidget(_divider())
        lay.addWidget(_sec("GESTURE BINDINGS"))
        for gest, action in (
            ("Fist (hold 0.6s)",   "Lock screen"),
            ("Swipe left ←",       "Alt+Shift+Tab / Ctrl+Shift+Tab"),
            ("Swipe right →",      "Alt+Tab / Ctrl+Tab"),
            ("Thumbs up",          "Volume +"),
            ("Thumbs down",        "Volume −"),
            ("Open palm hold",     "Play / Pause"),
        ):
            rw = QWidget()
            rl = QHBoxLayout(rw); rl.setContentsMargins(0, 2, 0, 2)
            n = QLabel(gest); n.setFont(inter(11, 500))
            n.setStyleSheet(f"color:{J.TEXT_PRI.name()};")
            rl.addWidget(n); rl.addStretch(1)
            h = QLabel(action); h.setFont(mono(9, 400))
            h.setStyleSheet(f"color:{J.GREEN.name()};")
            rl.addWidget(h)
            lay.addWidget(rw)

        # Recent gesture log
        lay.addWidget(_divider())
        lay.addWidget(_sec("RECENT EVENTS"))
        self._log_label = QLabel("(no gestures yet)")
        self._log_label.setFont(mono(10, 400))
        self._log_label.setStyleSheet(f"color:{J.TEXT_MUT.name()};")
        self._log_label.setWordWrap(True)
        lay.addWidget(self._log_label)

        # Voice trigger hints
        lay.addWidget(_divider())
        lay.addWidget(_sec("TRY SAYING"))
        for cmd in (
            '"gesture mode on"',
            '"what am I holding"',
            '"what do you see"',
            '"snap a photo"',
        ):
            l = QLabel(cmd); l.setFont(mono(10, 400))
            l.setStyleSheet(f"color:{J.CYAN.name()};")
            lay.addWidget(l)

        lay.addStretch(1)
        root.addWidget(_scrollable(inner), 1)

        # Poll engine state every 800ms — light enough to never lag the GUI.
        self._poll = QTimer(self); self._poll.timeout.connect(self._refresh)
        self._poll.start(800)

        self._events: list[str] = []

    def _refresh(self) -> None:
        try:
            from core.gesture_engine import get_gesture_engine
            from core.vision_engine import get_vision_engine
            ge = get_gesture_engine()
            ve = get_vision_engine()
        except Exception:
            return

        cam_on = ve.is_running()
        ges_on = ge.is_enabled()
        self._cam_row.findChildren(QLabel)[1].setText("LIVE" if cam_on else "OFFLINE")
        self._cam_row.findChildren(QLabel)[1].setStyleSheet(
            f"color:{(J.GREEN if cam_on else J.TEXT_MUT).name()};"
        )
        self._gest_row.findChildren(QLabel)[1].setText("ACTIVE" if ges_on else "OFF")
        self._gest_row.findChildren(QLabel)[1].setStyleSheet(
            f"color:{(J.GREEN if ges_on else J.TEXT_MUT).name()};"
        )

        # Listen for emitted gestures.
        if not getattr(self, "_listener_attached", False):
            try:
                ge.add_listener(self._on_gesture)
                self._listener_attached = True
            except Exception:
                pass

    def _on_gesture(self, name: str) -> None:
        ts = time.strftime("%H:%M:%S")
        self._events.insert(0, f"[{ts}] {name}")
        del self._events[8:]
        self._log_label.setText("\n".join(self._events))


# ─── Workbook panel ───────────────────────────────────────────── #

class WorkbookPanel(_GlassPanel):
    """Quick-action surface for the expense / tasks / meetings workbook."""

    def __init__(self, parent=None):
        super().__init__(parent)
        root = QVBoxLayout(self); root.setContentsMargins(0, 0, 0, 0); root.setSpacing(0)
        root.addWidget(_PanelHeader("WORKBOOK", "Expenses · tasks · cloud sync", J.AMBER))

        inner = QWidget()
        lay = QVBoxLayout(inner); lay.setContentsMargins(16, 18, 16, 18); lay.setSpacing(16)

        # KPI tiles
        lay.addWidget(_sec("THIS MONTH"))
        grid = QGridLayout(); grid.setSpacing(8)
        self._spend_tile = _StatCard("—",  "Total\nSpend",     J.AMBER)
        self._top_tile   = _StatCard("—",  "Top\nCategory",    J.MAGENTA)
        self._tasks_tile = _StatCard("—",  "Open\nTasks",      J.CYAN)
        self._search_tile = _StatCard("—", "Cached\nSearches", J.PURPLE)
        grid.addWidget(self._spend_tile, 0, 0)
        grid.addWidget(self._top_tile,   0, 1)
        grid.addWidget(self._tasks_tile, 1, 0)
        grid.addWidget(self._search_tile,1, 1)
        lay.addLayout(grid)

        # Voice trigger hints
        lay.addWidget(_divider())
        lay.addWidget(_sec("VOICE TRIGGERS"))
        for cmd in (
            '"500 rupees food pe kharch kiye"',
            '"is mahine kitna kharcha"',
            '"add task finish report by friday"',
            '"open expense sheet"',
            '"sync to google sheets"',
            '"search online what is python"',
        ):
            l = QLabel(cmd); l.setFont(mono(10, 400))
            l.setStyleSheet(f"color:{J.CYAN.name()};")
            lay.addWidget(l)

        # File path
        lay.addWidget(_divider())
        lay.addWidget(_sec("WORKBOOK FILE"))
        self._path_label = QLabel("data/jarvis_workbook.xlsx")
        self._path_label.setFont(mono(9, 400))
        self._path_label.setStyleSheet(f"color:{J.TEXT_MUT.name()};")
        self._path_label.setWordWrap(True)
        lay.addWidget(self._path_label)

        lay.addStretch(1)
        root.addWidget(_scrollable(inner), 1)

        self._poll = QTimer(self); self._poll.timeout.connect(self._refresh)
        self._poll.start(2500)
        self._refresh()

    def _refresh(self) -> None:
        try:
            import openpyxl
            from pathlib import Path
            wb_path = Path(__file__).resolve().parents[2] / "data" / "jarvis_workbook.xlsx"
            if not wb_path.exists():
                self._spend_tile.findChildren(QLabel)[0].setText("—")
                return
            try:
                wb = openpyxl.load_workbook(wb_path, read_only=True, data_only=False)
            except Exception:
                return
            # This-month total + top category from Category Summary tab.
            try:
                cs = wb["Category Summary"]
                total, top_cat, top_v = 0.0, "—", -1.0
                for row in cs.iter_rows(min_row=2, values_only=True):
                    if not row or row[0] in (None, ""): continue
                    v = row[1] if isinstance(row[1], (int, float)) else 0.0
                    total += v
                    if v > top_v:
                        top_v, top_cat = v, str(row[0])
                self._spend_tile.findChildren(QLabel)[0].setText(f"₹{int(total)}")
                self._top_tile.findChildren(QLabel)[0].setText(
                    top_cat if top_cat != "—" and len(top_cat) <= 12
                    else (top_cat[:11] + "…" if top_cat != "—" else "—")
                )
            except Exception:
                pass
            try:
                tasks = wb["Tasks"]
                open_n = 0
                for row in tasks.iter_rows(min_row=2, values_only=True):
                    if row and (row[4] or "").lower() == "open":
                        open_n += 1
                self._tasks_tile.findChildren(QLabel)[0].setText(str(open_n))
            except Exception:
                pass
            wb.close()
        except Exception:
            pass

        # Cached searches count.
        try:
            from core.knowledge_cache import get_default_cache
            n = get_default_cache().stats().get("entries", 0)
            self._search_tile.findChildren(QLabel)[0].setText(str(n))
        except Exception:
            pass


# ─── Routines panel ───────────────────────────────────────────── #

class _RoutineCard(QFrame):
    """One routine. Header (name + toggle), trigger pill, action chips,
    last-fired footer, run-now button.

    Signals via parent callbacks:
      on_toggle(routine_id, enabled_now)
      on_run(routine_id)
    """

    def __init__(self, routine, *, on_toggle, on_run, parent=None):
        super().__init__(parent)
        self._r = routine
        self._on_toggle = on_toggle
        self._on_run = on_run
        self._color = J.CYAN if routine.enabled else J.TEXT_MUT
        self.setObjectName("RoutineCard")
        self.setStyleSheet(
            "QFrame#RoutineCard{"
            f"background:rgba({J.BG_ELE.red()},{J.BG_ELE.green()},{J.BG_ELE.blue()},0.65);"
            f"border:1px solid rgba({self._color.red()},{self._color.green()},{self._color.blue()},0.32);"
            "border-radius:10px;"
            "}"
        )
        lay = QVBoxLayout(self)
        lay.setContentsMargins(12, 10, 12, 10); lay.setSpacing(6)

        # ── Header row ─────────────────────────────────────────
        hr = QHBoxLayout(); hr.setSpacing(8)
        title = QLabel(routine.name); title.setFont(inter(11, 700))
        title.setStyleSheet(f"color:{J.TEXT_PRI.name()};letter-spacing:0.5px;")
        hr.addWidget(title, 1)

        toggle = _ToggleChip(routine.enabled, parent=self)
        toggle.clicked.connect(self._handle_toggle)
        self._toggle = toggle
        hr.addWidget(toggle)
        lay.addLayout(hr)

        # ── Description ────────────────────────────────────────
        if routine.description:
            desc = QLabel(routine.description); desc.setFont(mono(9, 400))
            desc.setStyleSheet(f"color:{J.TEXT_MUT.name()};")
            desc.setWordWrap(True)
            lay.addWidget(desc)

        # ── Trigger pill ───────────────────────────────────────
        trig = QLabel(routine.trigger.summary()); trig.setFont(mono(9, 700))
        trig.setStyleSheet(
            f"color:{self._color.name()};"
            f"background:rgba({self._color.red()},{self._color.green()},{self._color.blue()},0.15);"
            "padding:3px 8px;border-radius:6px;"
        )
        trig.setAlignment(Qt.AlignLeft | Qt.AlignVCenter)
        wrap = QHBoxLayout(); wrap.setContentsMargins(0,2,0,2); wrap.setSpacing(6)
        wrap.addWidget(trig); wrap.addStretch(1)
        lay.addLayout(wrap)

        # ── Action chips ───────────────────────────────────────
        for a in routine.actions[:6]:
            row = QHBoxLayout(); row.setContentsMargins(0,0,0,0); row.setSpacing(6)
            dot = QLabel("•"); dot.setFont(mono(11, 700))
            dot.setStyleSheet(f"color:{J.GREEN.name()};")
            dot.setFixedWidth(10)
            row.addWidget(dot)
            txt = QLabel(a.summary()); txt.setFont(mono(9, 400))
            txt.setStyleSheet(f"color:{J.TEXT_SEC.name()};")
            row.addWidget(txt, 1)
            lay.addLayout(row)
        if len(routine.actions) > 6:
            more = QLabel(f"  + {len(routine.actions) - 6} more"); more.setFont(mono(9, 400))
            more.setStyleSheet(f"color:{J.TEXT_MUT.name()};")
            lay.addWidget(more)

        # ── Footer (run-now + last fired) ──────────────────────
        ft = QHBoxLayout(); ft.setSpacing(6)
        run = QPushButton("Run now")
        run.setFont(mono(9, 700))
        run.setCursor(Qt.PointingHandCursor)
        run.setStyleSheet(
            f"QPushButton{{background:rgba({J.CYAN.red()},{J.CYAN.green()},{J.CYAN.blue()},0.15);"
            f"color:{J.CYAN.name()};border:1px solid rgba({J.CYAN.red()},{J.CYAN.green()},{J.CYAN.blue()},0.4);"
            "border-radius:5px;padding:4px 10px;}}"
            f"QPushButton:hover{{background:rgba({J.CYAN.red()},{J.CYAN.green()},{J.CYAN.blue()},0.30);}}"
        )
        run.clicked.connect(lambda: self._on_run(self._r.id))
        ft.addWidget(run)
        ft.addStretch(1)
        if routine.fire_count:
            from datetime import datetime as _dt
            last = _dt.fromtimestamp(routine.last_fired_ts).strftime("%b %d %H:%M")
            ft_lbl = QLabel(f"fired {routine.fire_count}× · last {last}")
        else:
            ft_lbl = QLabel("never fired")
        ft_lbl.setFont(mono(8, 400))
        ft_lbl.setStyleSheet(f"color:{J.TEXT_MUT.name()};")
        ft.addWidget(ft_lbl)
        lay.addLayout(ft)

    def _handle_toggle(self):
        new_state = not self._r.enabled
        self._toggle.set_on(new_state)
        self._r.enabled = new_state
        self._on_toggle(self._r.id, new_state)


class _ToggleChip(QPushButton):
    """Small ON/off pill button."""

    def __init__(self, on: bool, parent=None):
        super().__init__("ENABLED" if on else "DISABLED", parent)
        self._on = on
        self.setCursor(Qt.PointingHandCursor)
        self.setFont(mono(8, 700))
        self.setFixedHeight(20)
        self.set_on(on)

    def set_on(self, on: bool) -> None:
        self._on = on
        col = J.GREEN if on else J.TEXT_MUT
        self.setText("ENABLED" if on else "DISABLED")
        self.setStyleSheet(
            f"QPushButton{{background:rgba({col.red()},{col.green()},{col.blue()},0.15);"
            f"color:{col.name()};border:1px solid rgba({col.red()},{col.green()},{col.blue()},0.55);"
            "border-radius:10px;padding:1px 9px;letter-spacing:1px;}"
            f"QPushButton:hover{{background:rgba({col.red()},{col.green()},{col.blue()},0.28);}}"
        )


class RoutinesPanel(_GlassPanel):
    """Right-rail Automation tab. Lists routines, lets the user enable
    /disable / run them. Refresh button re-reads from disk.
    """

    def __init__(self, parent=None):
        super().__init__(parent)
        root = QVBoxLayout(self); root.setContentsMargins(0,0,0,0); root.setSpacing(0)
        root.addWidget(_PanelHeader("AUTOMATION",
                                    "Routines · triggers · workflows", J.PURPLE))

        # Toolbar row
        bar = QHBoxLayout(); bar.setContentsMargins(16, 12, 16, 0); bar.setSpacing(8)
        refresh = QPushButton("↻  Reload")
        refresh.setFont(mono(9, 700))
        refresh.setCursor(Qt.PointingHandCursor)
        refresh.setStyleSheet(
            f"QPushButton{{background:rgba({J.CYAN.red()},{J.CYAN.green()},{J.CYAN.blue()},0.12);"
            f"color:{J.CYAN.name()};border:1px solid rgba({J.CYAN.red()},{J.CYAN.green()},{J.CYAN.blue()},0.32);"
            "border-radius:5px;padding:4px 10px;}}"
        )
        refresh.clicked.connect(self.refresh)
        bar.addWidget(refresh)
        bar.addStretch(1)
        self._count = QLabel("")
        self._count.setFont(mono(9, 400))
        self._count.setStyleSheet(f"color:{J.TEXT_MUT.name()};")
        bar.addWidget(self._count)
        bar_wrap = QWidget(); bar_wrap.setLayout(bar)
        root.addWidget(bar_wrap)

        # Cards list
        self._inner = QWidget()
        self._lay = QVBoxLayout(self._inner)
        self._lay.setContentsMargins(16, 12, 16, 18); self._lay.setSpacing(10)
        self._empty_label = QLabel(
            "No routines found.\n\n"
            "Edit data/routines.json to add yours,\n"
            "then hit Reload."
        )
        self._empty_label.setFont(mono(9, 400))
        self._empty_label.setStyleSheet(f"color:{J.TEXT_MUT.name()};")
        self._empty_label.setAlignment(Qt.AlignCenter)
        self._empty_label.setWordWrap(True)
        self._lay.addWidget(self._empty_label)
        root.addWidget(_scrollable(self._inner), 1)

        # Auto-populate on construction.
        self.refresh()

    # ── public ──────────────────────────────────────────────────

    def refresh(self) -> None:
        """Rebuild every card from the current engine state."""
        try:
            from core.automation import get_engine
            routines = get_engine().list_routines()
        except Exception as e:
            routines = []
            self._empty_label.setText(f"Engine offline:\n{e}")

        # Clear current items.
        for i in reversed(range(self._lay.count())):
            item = self._lay.itemAt(i).widget()
            if item is not None and item is not self._empty_label:
                item.setParent(None)

        if not routines:
            self._empty_label.show()
            self._count.setText("0 routines")
            return

        self._empty_label.hide()
        # Move empty_label out of layout flow without removing — easier
        # than juggling parents. Insert cards before the (hidden) empty.
        idx = 0
        for r in routines:
            card = _RoutineCard(r,
                                on_toggle=self._on_toggle,
                                on_run=self._on_run)
            self._lay.insertWidget(idx, card)
            idx += 1
        self._count.setText(f"{len(routines)} routines")

    # ── callbacks ───────────────────────────────────────────────

    def _on_toggle(self, rid: str, enabled: bool) -> None:
        try:
            from core.automation import get_engine
            get_engine().set_enabled(rid, enabled)
        except Exception:
            pass

    def _on_run(self, rid: str) -> None:
        try:
            from core.automation import get_engine
            get_engine().run_routine(rid, reason="GUI run-now")
        except Exception:
            pass


# ─── Notifications panel ──────────────────────────────────────── #

_EVENT_ICONS = {
    "routine_fire":      ("RUN", J.PURPLE),
    "skill_reload":      ("RLD", J.CYAN),
    "workspace_changed": ("WKS", J.AMBER),
    "crash_recovered":   ("CRH", J.RED),
    "automation_started":("AUT", J.GREEN),
    "boot_start":        ("BUP", J.GREEN),
    "gesture":           ("GES", J.PURPLE),
}


class NotificationsPanel(_GlassPanel):
    """Persistent event center — last 200 events from log_setup.event().

    Subscribes once on construction; the subscriber lives until the
    panel widget is destroyed. New events arrive on the caller's thread
    (which could be any worker), so we marshal to the GUI thread via a
    queued Qt signal before mutating the list.
    """

    _event_arrived = pyqtSignal(dict)

    def __init__(self, parent=None):
        super().__init__(parent)
        root = QVBoxLayout(self); root.setContentsMargins(0,0,0,0); root.setSpacing(0)
        root.addWidget(_PanelHeader("EVENTS",
                                    "Live activity · last 200", J.CYAN))

        bar = QHBoxLayout(); bar.setContentsMargins(16, 10, 16, 4); bar.setSpacing(8)
        self._filter = QComboBox()
        self._filter.addItems(["All", "Routines", "Gestures", "Workspace",
                               "Crashes", "Boot"])
        self._filter.setFixedHeight(24)
        self._filter.setStyleSheet(
            f"QComboBox{{background:rgba({J.BG_ELE.red()},{J.BG_ELE.green()},{J.BG_ELE.blue()},0.65);"
            f"color:{J.TEXT_PRI.name()};border:1px solid rgba({J.BORDER.red()},{J.BORDER.green()},{J.BORDER.blue()},0.5);"
            "border-radius:5px;padding:2px 8px;}"
            "QComboBox QAbstractItemView{background:#0d1525;color:#dcecf0;}"
        )
        self._filter.currentTextChanged.connect(self._render)
        bar.addWidget(self._filter)
        bar.addStretch(1)
        self._count = QLabel("0 events")
        self._count.setFont(mono(9, 400))
        self._count.setStyleSheet(f"color:{J.TEXT_MUT.name()};")
        bar.addWidget(self._count)
        bar_wrap = QWidget(); bar_wrap.setLayout(bar)
        root.addWidget(bar_wrap)

        self._inner = QWidget()
        self._lay = QVBoxLayout(self._inner)
        self._lay.setContentsMargins(16, 6, 16, 18); self._lay.setSpacing(4)
        self._lay.addStretch(1)
        root.addWidget(_scrollable(self._inner), 1)

        self._events: list[dict] = []
        # Backfill with whatever happened before the panel was constructed.
        try:
            from core.log_setup import snapshot_events, subscribe_events
            self._events = snapshot_events()[-200:]
            self._event_arrived.connect(self._on_event_qt_thread)
            self._unsub = subscribe_events(self._event_arrived.emit)
        except Exception:
            self._unsub = None
        self._render()

    # ── arrival ────────────────────────────────────────────────────

    def _on_event_qt_thread(self, payload: dict) -> None:
        self._events.append(payload)
        # Bounded — match the source ring buffer.
        if len(self._events) > 200:
            self._events = self._events[-200:]
        self._render()

    # ── render ─────────────────────────────────────────────────────

    def _matches_filter(self, ev: dict) -> bool:
        f = self._filter.currentText()
        if f == "All":
            return True
        kind = ev.get("event", "")
        mapping = {
            "Routines":  ("routine_fire", "automation_started"),
            "Gestures":  ("gesture",),
            "Workspace": ("workspace_changed",),
            "Crashes":   ("crash_recovered", "crash_dump"),
            "Boot":      ("boot_start", "automation_started"),
        }
        return kind in mapping.get(f, ())

    def _render(self) -> None:
        # Clear current rows (keep the trailing stretch).
        for i in reversed(range(self._lay.count() - 1)):
            w = self._lay.itemAt(i).widget()
            if w is not None:
                w.setParent(None)

        shown = [e for e in reversed(self._events) if self._matches_filter(e)]
        for ev in shown[:120]:
            self._lay.insertWidget(self._lay.count() - 1, _EventRow(ev))
        self._count.setText(f"{len(shown)} / {len(self._events)} events")


class _EventRow(QWidget):
    def __init__(self, ev: dict, parent=None):
        super().__init__(parent)
        lay = QHBoxLayout(self); lay.setContentsMargins(0, 2, 0, 2); lay.setSpacing(8)
        from datetime import datetime as _dt
        ts = _dt.fromtimestamp(ev.get("ts", 0)).strftime("%H:%M:%S")
        tlabel = QLabel(ts); tlabel.setFont(mono(8, 400))
        tlabel.setStyleSheet(f"color:{J.TEXT_MUT.name()};")
        tlabel.setFixedWidth(56)
        lay.addWidget(tlabel)

        kind = ev.get("event", "?")
        tag_txt, tag_col = _EVENT_ICONS.get(kind, ("EVT", J.TEXT_MUT))
        tag = QLabel(tag_txt); tag.setFont(mono(8, 700))
        tag.setStyleSheet(
            f"color:{tag_col.name()};"
            f"background:rgba({tag_col.red()},{tag_col.green()},{tag_col.blue()},0.15);"
            "padding:1px 6px;border-radius:4px;"
        )
        tag.setFixedWidth(36)
        tag.setAlignment(Qt.AlignCenter)
        lay.addWidget(tag)

        # Summarise non-ts/non-event fields into one line.
        rest = {k: v for k, v in ev.items() if k not in ("ts", "event")}
        body = " · ".join(f"{k}={v}" for k, v in rest.items())
        if not body:
            body = kind
        else:
            body = f"{kind} · {body}"
        text = QLabel(body); text.setFont(mono(9, 400))
        text.setStyleSheet(f"color:{J.TEXT_SEC.name()};")
        text.setWordWrap(False)
        lay.addWidget(text, 1)


# ─── Smart suggestions panel ──────────────────────────────────── #

class _SuggestionCard(QPushButton):
    """One context-aware action card. Click → emits its action.

    Three flavors via ``kind``:
        skill    — runs a registered skill by name
        routine  — fires a routine by id
        text     — submits the text to the chat input (free-form)
    """

    def __init__(self, *, title: str, sub: str, kind: str, payload: str,
                 color, on_click, parent=None):
        super().__init__(parent)
        self._kind = kind
        self._payload = payload
        self._on_click = on_click
        self.setCursor(Qt.PointingHandCursor)
        self.setFixedHeight(64)
        self.setFont(inter(11, 600))
        col = color
        self.setStyleSheet(
            f"QPushButton{{text-align:left;padding:8px 12px;"
            f"background:rgba({col.red()},{col.green()},{col.blue()},0.10);"
            f"color:{J.TEXT_PRI.name()};"
            f"border:1px solid rgba({col.red()},{col.green()},{col.blue()},0.30);"
            "border-radius:8px;}}"
            f"QPushButton:hover{{background:rgba({col.red()},{col.green()},{col.blue()},0.22);"
            f"border-color:rgba({col.red()},{col.green()},{col.blue()},0.60);}}"
        )
        self.setText(f"  {title}\n  {sub}")
        self.clicked.connect(self._handle)

    def _handle(self):
        self._on_click(self._kind, self._payload)


class SmartSuggestionsPanel(_GlassPanel):
    """Context-aware quick-action grid.

    Refreshes its cards based on the *current* workspace profile, the
    time of day, battery state, and the routines currently enabled.
    Updates on ``tick_slow`` (~15 Hz, throttled to once per 2 s
    internally) and on every workspace transition.

    Card destinations:
      * ``skill`` payloads → run the named skill via the breaker
      * ``routine`` payloads → fire the routine by id
      * ``text`` payloads → submit raw text to the brain (via signal)
    """

    submit_text = pyqtSignal(str)

    def __init__(self, parent=None):
        super().__init__(parent)
        self._bus = get_bus()
        self._bus.tick_slow.connect(self._maybe_refresh)
        self._last_refresh = 0.0
        self._last_signature = ""

        root = QVBoxLayout(self); root.setContentsMargins(0,0,0,0); root.setSpacing(0)
        root.addWidget(_PanelHeader("QUICK ACTIONS",
                                    "Context-aware shortcuts", J.AMBER))

        self._ctx_label = QLabel("Context: detecting...")
        self._ctx_label.setFont(mono(9, 700))
        self._ctx_label.setStyleSheet(f"color:{J.AMBER.name()};letter-spacing:1px;")
        self._ctx_label.setContentsMargins(16, 10, 16, 6)
        root.addWidget(self._ctx_label)

        self._grid_host = QWidget()
        self._grid = QGridLayout(self._grid_host)
        self._grid.setContentsMargins(16, 4, 16, 14); self._grid.setSpacing(8)
        root.addWidget(_scrollable(self._grid_host), 1)

        # Wire workspace transitions for immediate refresh (instead of
        # waiting for the next tick_slow).
        try:
            from core.workspace_profile import get_monitor as _gw
            _gw().subscribe(lambda *_: QTimer.singleShot(0, self._refresh))
        except Exception:
            pass

        QTimer.singleShot(120, self._refresh)

    # ── throttled tick ─────────────────────────────────────────────

    def _maybe_refresh(self):
        now = time.monotonic()
        if now - self._last_refresh < 2.0:
            return
        self._refresh()

    # ── card selection ─────────────────────────────────────────────

    def _current_context(self) -> dict:
        """Snapshot of everything the suggestion picker considers."""
        ctx = {"workspace": "IDLE", "hour": time.localtime().tm_hour,
               "battery": 100, "on_battery": False, "is_idle": False}
        try:
            from core.workspace_profile import get_monitor as _gw
            ctx["workspace"] = _gw().snapshot().profile
        except Exception:
            pass
        try:
            from core.power_state import get_monitor as _gp
            snap = _gp().snapshot()
            ctx["battery"] = snap.percent
            ctx["on_battery"] = snap.on_battery
            ctx["is_idle"] = snap.is_idle
        except Exception:
            pass
        return ctx

    def _pick_suggestions(self, ctx: dict) -> list:
        """Return a list of dicts: title/sub/kind/payload/color.

        Each rule below decides whether to surface its card based on the
        context. Earlier rules take priority (we cap at 9 cards). The
        ordering deliberately prioritises *actionable now* over generic.
        """
        s: list = []
        ws = ctx["workspace"]
        hour = ctx["hour"]
        battery = ctx["battery"]
        on_battery = ctx["on_battery"]

        # ── critical (battery low) ──
        if on_battery and battery <= 25:
            s.append({"title": "Battery saver", "sub": f"On battery, {battery}%",
                      "kind": "routine", "payload": "battery_saver",
                      "color": J.RED})

        # ── workspace-driven ──
        if ws == "CODING":
            s.append({"title": "Pomodoro · 4×", "sub": "25/5 focus blocks",
                      "kind": "routine", "payload": "pomodoro", "color": J.MAGENTA})
            s.append({"title": "Deep focus on", "sub": "Block distractions + DND",
                      "kind": "skill", "payload": "focus_deep_on", "color": J.PURPLE})
            s.append({"title": "Snap window left",  "sub": "Hands-free split",
                      "kind": "skill", "payload": "window_snap_left", "color": J.CYAN})
            s.append({"title": "System stats", "sub": "CPU · RAM · disk",
                      "kind": "skill", "payload": "system_stats", "color": J.GREEN})
        elif ws == "MEETING":
            s.append({"title": "Volume 15%", "sub": "Quiet background",
                      "kind": "skill", "payload": "volume_set",
                      "color": J.MAGENTA})
            s.append({"title": "Gesture mode on", "sub": "Pinch to mute mic",
                      "kind": "skill", "payload": "gesture_mode_on",
                      "color": J.PURPLE})
            s.append({"title": "Note this meeting", "sub": "Append to scratch.md",
                      "kind": "text", "payload": "note this: meeting notes — ",
                      "color": J.AMBER})
        elif ws == "MEDIA":
            s.append({"title": "Volume up",   "sub": "+10%",
                      "kind": "skill", "payload": "volume_up",   "color": J.CYAN})
            s.append({"title": "Volume down", "sub": "-10%",
                      "kind": "skill", "payload": "volume_down", "color": J.CYAN})
        elif ws == "GAMING":
            s.append({"title": "Show desktop", "sub": "Min everything",
                      "kind": "skill", "payload": "show_desktop", "color": J.AMBER})
        else:
            s.append({"title": "Daily summary", "sub": "stats → AI brief",
                      "kind": "routine", "payload": "daily_summary_pipeline",
                      "color": J.PURPLE})

        # ── time-of-day ──
        if 6 <= hour < 11:
            s.append({"title": "Good morning", "sub": "Boot workspace",
                      "kind": "routine", "payload": "smart_morning",
                      "color": J.GREEN})
        elif 18 <= hour < 22:
            s.append({"title": "Coding mode", "sub": "Evening setup",
                      "kind": "routine", "payload": "coding_mode",
                      "color": J.CYAN})
        elif hour >= 22 or hour < 6:
            s.append({"title": "Wind down", "sub": "Dim · quiet · DND",
                      "kind": "routine", "payload": "evening_wind_down",
                      "color": J.PURPLE})

        # ── always-useful (fill to ~9) ──
        defaults = [
            ("Lock screen", "Win+L equivalent",  "skill",   "lock_screen",   J.RED),
            ("Volume mute", "Toggle audio",      "skill",   "volume_mute",   J.MAGENTA),
            ("Screen snip", "Win+Shift+S",       "skill",   "show_desktop",  J.CYAN),
            ("File search", "Find a doc",        "text",    "find file ",    J.AMBER),
            ("List routines", "Show all",        "skill",   "routine_list",  J.GREEN),
        ]
        for title, sub, kind, payload, col in defaults:
            if len(s) >= 9:
                break
            if not any(c.get("payload") == payload for c in s):
                s.append({"title": title, "sub": sub, "kind": kind,
                          "payload": payload, "color": col})

        return s[:9]

    # ── render ─────────────────────────────────────────────────────

    def _refresh(self):
        self._last_refresh = time.monotonic()
        ctx = self._current_context()
        # Skip rebuild if context fingerprint hasn't changed.
        sig = f"{ctx['workspace']}|{ctx['hour']}|{ctx['battery']//5}|{ctx['on_battery']}"
        if sig == self._last_signature:
            return
        self._last_signature = sig

        # Update header
        parts = [f"workspace={ctx['workspace'].lower()}"]
        parts.append(f"{ctx['battery']}% {'bat' if ctx['on_battery'] else 'AC'}")
        if ctx['is_idle']:
            parts.append("idle")
        parts.append(f"{ctx['hour']:02d}h")
        self._ctx_label.setText("  ·  ".join(parts))

        # Clear grid
        while self._grid.count():
            item = self._grid.takeAt(0)
            w = item.widget() if item is not None else None
            if w is not None:
                w.setParent(None)

        cards = self._pick_suggestions(ctx)
        for i, c in enumerate(cards):
            card = _SuggestionCard(
                title=c["title"], sub=c["sub"], kind=c["kind"],
                payload=c["payload"], color=c["color"],
                on_click=self._on_card,
            )
            row, col = divmod(i, 2)
            self._grid.addWidget(card, row, col)

    def _on_card(self, kind: str, payload: str) -> None:
        if kind == "skill":
            try:
                from core.skill_breaker import call
                from core.skill_registry import REGISTRY
                s = REGISTRY.get(payload)
                if s is not None:
                    call(payload, s.run, {})
            except Exception:
                pass
        elif kind == "routine":
            try:
                from core.automation import get_engine
                get_engine().run_routine(payload, reason="quick-action")
            except Exception:
                pass
        elif kind == "text":
            self.submit_text.emit(payload)


# ─── Right-panel stack ────────────────────────────────────────── #

_TAB_INDEX = {
    "home":     0,
    "chat":     1,
    "auto":     1,   # automation chips shown inside GlassChatPanel
    "brain":    2,
    "memory":   3,
    "system":   4,
    "settings": 5,
    "vision":   6,
    "workbook": 7,
    "routines": 8,
    "events":   9,
    "quick":    10,
}


class RightPanelStack(QWidget):
    """Hosts all tab panels; exposes `chat_panel` and `switch_tab(key)`."""

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setFixedWidth(390)

        from .glass_chat_panel import GlassChatPanel
        self.chat_panel = GlassChatPanel()

        self._stack = QStackedWidget(self)
        self._stack.setFixedWidth(390)
        self._stack.addWidget(HomePanel())         # 0
        self._stack.addWidget(self.chat_panel)     # 1
        self._stack.addWidget(BrainPanel())        # 2
        self._stack.addWidget(MemoryPanel())       # 3
        self._stack.addWidget(SystemPanel())       # 4
        self._stack.addWidget(SettingsPanel())     # 5
        self._stack.addWidget(VisionPanel())       # 6
        self._stack.addWidget(WorkbookPanel())     # 7
        self.routines_panel = RoutinesPanel()
        self._stack.addWidget(self.routines_panel) # 8
        self.notifications_panel = NotificationsPanel()
        self._stack.addWidget(self.notifications_panel)  # 9
        self.quick_panel = SmartSuggestionsPanel()
        self._stack.addWidget(self.quick_panel)          # 10
        self._stack.setCurrentIndex(1)

        lay = QVBoxLayout(self)
        lay.setContentsMargins(0, 0, 0, 0); lay.setSpacing(0)
        lay.addWidget(self._stack)

    def switch_tab(self, key: str) -> None:
        self._stack.setCurrentIndex(_TAB_INDEX.get(key, 1))
        self.chat_panel.set_automation_visible(key == "auto")
