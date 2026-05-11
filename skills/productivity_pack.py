"""Productivity pack: daily briefing, focus mode, smart paste.

Daily briefing  → "good morning" / "daily briefing"
Focus mode      → "focus mode on" — closes distractions, opens work apps,
                  starts a pomodoro timer (configurable in settings.json)
Smart paste     → "smart paste" — reformats the current clipboard:
                  long text → summary, JSON → pretty-print, URL → page title,
                  code → trimmed indentation.
"""

from __future__ import annotations

import json
import logging
import os
import re
import subprocess
import sys
import threading
import time
from datetime import datetime
from typing import Optional
from urllib.parse import urlparse

_HERE = os.path.dirname(os.path.abspath(__file__))
_ROOT = os.path.dirname(_HERE)
if _ROOT not in sys.path:
    sys.path.insert(0, _ROOT)

from core import settings  # noqa: E402
from core.skill_registry import skill  # noqa: E402

log = logging.getLogger(__name__)


# --------------------------------------------------------------------------- #
#  Daily briefing                                                             #
# --------------------------------------------------------------------------- #

def _time_greeting() -> str:
    h = datetime.now().hour
    if h < 5: return "Late night still up"
    if h < 12: return "Good morning"
    if h < 17: return "Good afternoon"
    if h < 21: return "Good evening"
    return "Good night"


def _battery_line() -> str:
    try:
        import psutil
        b = psutil.sensors_battery()
        if not b: return ""
        return f"Battery {b.percent:.0f}% ({'charging' if b.power_plugged else 'on battery'})"
    except Exception:
        return ""


def _weather_line() -> str:
    """Best-effort weather using wttr.in — no API key needed."""
    try:
        import urllib.request
        loc = "Delhi"
        try:
            from core.memory import UserMemory
            m = UserMemory(os.path.join(_ROOT, "data", "user_memory.json"))
            loc = m.get("location") or loc
        except Exception:
            pass
        url = f"https://wttr.in/{loc}?format=%C+%t&m"
        with urllib.request.urlopen(url, timeout=3) as r:
            txt = r.read().decode("utf-8").strip()
        if txt and "Sorry" not in txt:
            return f"Weather in {loc}: {txt}"
    except Exception:
        pass
    return ""


def _workbook_lines() -> list[str]:
    out = []
    try:
        import openpyxl
        path = os.path.join(_ROOT, "data", "jarvis_workbook.xlsx")
        if not os.path.exists(path):
            return out
        wb = openpyxl.load_workbook(path, read_only=True, data_only=False)
        if "Tasks" in wb.sheetnames:
            tasks = wb["Tasks"]
            open_tasks = []
            for row in tasks.iter_rows(min_row=2, values_only=True):
                if row and (row[4] or "").lower() == "open":
                    open_tasks.append(str(row[1]))
            if open_tasks:
                top = open_tasks[:3]
                out.append(f"Top open tasks: {' · '.join(top)}")
        if "Category Summary" in wb.sheetnames:
            cs = wb["Category Summary"]
            total = 0.0
            for row in cs.iter_rows(min_row=2, values_only=True):
                v = row[1] if (row and isinstance(row[1], (int, float))) else 0
                total += v
            if total:
                out.append(f"Spend this month: ₹{int(total)}")
        wb.close()
    except Exception:
        pass
    return out


def _reminders_line() -> str:
    try:
        from core.scheduler import ReminderScheduler  # noqa: F401
    except Exception:
        return ""
    return ""


@skill(
    name="daily_briefing",
    description="Speak a quick personalized briefing: greeting, weather, tasks, expenses, system status.",
    patterns=[
        "good morning", "daily briefing", "morning briefing",
        "briefing do", "aaj ka briefing", "mujhe briefing do",
        "good morning aeris", "good morning jarvis",
        "morning update", "start of day report",
        "mera daily briefing batao",
    ],
    required_entities=[],
)
def daily_briefing(slots: dict) -> str:
    name = settings.assistant_name()
    user_name = ""
    try:
        from core.memory import UserMemory
        m = UserMemory(os.path.join(_ROOT, "data", "user_memory.json"))
        user_name = m.get("name") or ""
    except Exception:
        pass

    lines = []
    greet = _time_greeting()
    if user_name:
        lines.append(f"{greet}, {user_name}. {name} reporting in.")
    else:
        lines.append(f"{greet}, sir. {name} reporting in.")
    lines.append(f"Today is {datetime.now().strftime('%A, %d %B %Y')}.")

    w = _weather_line()
    if w: lines.append(w)
    b = _battery_line()
    if b: lines.append(b)
    for w_line in _workbook_lines():
        lines.append(w_line)
    try:
        from core.knowledge_cache import get_default_cache
        n = get_default_cache().stats().get("entries", 0)
        if n:
            lines.append(f"{n} searches cached locally.")
    except Exception:
        pass
    lines.append("Have a productive day, sir.")
    return "\n".join(lines)


# --------------------------------------------------------------------------- #
#  Focus mode                                                                 #
# --------------------------------------------------------------------------- #

_FOCUS_STATE: dict = {"on": False, "since": None, "timer_thread": None,
                      "timer_cancel": None}


def _close_apps(names: list[str]) -> list[str]:
    closed = []
    try:
        import psutil
    except Exception:
        return closed
    name_set = {n.lower() for n in names}
    for proc in psutil.process_iter(["name"]):
        try:
            n = (proc.info["name"] or "").lower()
        except Exception:
            continue
        if any(n.startswith(target) or target in n for target in name_set):
            try:
                proc.kill()
                closed.append(proc.info["name"])
            except Exception:
                pass
    return closed


def _open_apps(names: list[str]) -> list[str]:
    opened = []
    import shutil
    for n in names:
        try:
            exe = shutil.which(n)
            if exe:
                subprocess.Popen([exe], close_fds=True)
                opened.append(n)
            else:
                subprocess.Popen(f"start {n}", shell=True)
                opened.append(n)
        except Exception:
            pass
    return opened


def _start_pomodoro(minutes: int) -> None:
    if _FOCUS_STATE.get("timer_thread"):
        c = _FOCUS_STATE.get("timer_cancel")
        if c: c.set()
    cancel = threading.Event()
    def _w():
        if cancel.wait(minutes * 60):
            return
        try:
            from win10toast import ToastNotifier
            ToastNotifier().show_toast(
                f"{settings.assistant_name()} — Focus session done",
                f"{minutes} min complete. Take a break.",
                duration=8, threaded=True,
            )
        except Exception:
            import ctypes
            try: ctypes.windll.user32.MessageBeep(0xFFFFFFFF)
            except Exception: pass
    t = threading.Thread(target=_w, daemon=True, name=f"FocusPomodoro-{minutes}")
    _FOCUS_STATE["timer_thread"] = t
    _FOCUS_STATE["timer_cancel"] = cancel
    t.start()


@skill(
    name="focus_mode_on",
    description="Enter focus mode: close distracting apps, open work apps, start a pomodoro.",
    patterns=[
        "focus mode on", "focus mode start karo",
        "deep work mode chalu karo", "kaam shuru mode",
        "study mode on", "work mode on",
        "distractions hatao",
        "let me focus", "no distractions please",
    ],
    required_entities=[],
)
def focus_mode_on(slots: dict) -> str:
    close = settings.get("focus_mode_close") or ["discord", "slack", "telegram"]
    open_ = settings.get("focus_mode_open") or ["code"]
    pomodoro = int(settings.get("focus_mode_pomodoro_min") or 25)

    closed = _close_apps(close)
    opened = _open_apps(open_)
    _start_pomodoro(pomodoro)
    _FOCUS_STATE["on"] = True
    _FOCUS_STATE["since"] = datetime.now()

    bits = [f"Focus mode ON ({pomodoro} min pomodoro)."]
    if closed: bits.append(f"Closed: {', '.join(closed[:5])}")
    if opened: bits.append(f"Opened: {', '.join(opened)}")
    return " · ".join(bits)


@skill(
    name="focus_mode_off",
    description="Exit focus mode and cancel the pomodoro.",
    patterns=[
        "focus mode off", "focus mode band karo",
        "focus mode stop karo", "deep work band karo",
        "exit focus mode", "stop focus mode",
    ],
    required_entities=[],
)
def focus_mode_off(slots: dict) -> str:
    if not _FOCUS_STATE.get("on"):
        return "Focus mode already off hai."
    c = _FOCUS_STATE.get("timer_cancel")
    if c: c.set()
    since = _FOCUS_STATE.get("since")
    dur = ""
    if since:
        m = int((datetime.now() - since).total_seconds() // 60)
        dur = f" ({m} min session)"
    _FOCUS_STATE["on"] = False
    _FOCUS_STATE["since"] = None
    _FOCUS_STATE["timer_thread"] = None
    _FOCUS_STATE["timer_cancel"] = None
    return f"Focus mode OFF{dur}. Well done, sir."


# --------------------------------------------------------------------------- #
#  Smart paste                                                                 #
# --------------------------------------------------------------------------- #

def _is_json(s: str) -> bool:
    s = (s or "").strip()
    if not s or s[0] not in "{[":
        return False
    try:
        json.loads(s); return True
    except Exception:
        return False


def _is_url(s: str) -> bool:
    s = (s or "").strip()
    if not re.match(r"^https?://", s, re.IGNORECASE):
        return False
    try:
        return bool(urlparse(s).netloc)
    except Exception:
        return False


def _fetch_title(url: str) -> Optional[str]:
    try:
        import requests
        from bs4 import BeautifulSoup
        r = requests.get(url, timeout=4,
                         headers={"User-Agent": "Mozilla/5.0 AERIS/1.0"})
        if r.status_code != 200:
            return None
        soup = BeautifulSoup(r.text, "html.parser")
        if soup.title and soup.title.string:
            return soup.title.string.strip()
    except Exception:
        return None
    return None


def _trim_code(s: str) -> str:
    """Remove a leading common indent (deindents pasted code)."""
    lines = s.split("\n")
    non_empty = [l for l in lines if l.strip()]
    if not non_empty:
        return s
    indents = [len(l) - len(l.lstrip(" ")) for l in non_empty]
    common = min(indents) if indents else 0
    if common == 0:
        return s
    return "\n".join(l[common:] if len(l) >= common else l for l in lines)


@skill(
    name="smart_paste",
    description="Reformat the current clipboard content: JSON → pretty, URL → page title, long text → summary, code → de-indent.",
    patterns=[
        "smart paste", "clipboard format karo",
        "format my clipboard", "clipboard ko clean karo",
        "clipboard summarize karo", "clipboard reformat karo",
        "auto format clipboard",
    ],
    required_entities=[],
)
def smart_paste(slots: dict) -> str:
    try:
        import pyperclip
    except Exception:
        return "pyperclip install nahi hai."
    try:
        content = pyperclip.paste() or ""
    except Exception as e:
        return f"Clipboard read fail: {e}"
    if not content.strip():
        return "Clipboard khali hai, sir."

    transformed: Optional[str] = None
    note = ""

    if _is_json(content):
        try:
            transformed = json.dumps(json.loads(content), indent=2, ensure_ascii=False)
            note = "JSON pretty-printed"
        except Exception:
            pass

    if transformed is None and _is_url(content):
        title = _fetch_title(content.strip())
        if title:
            transformed = f"{title}\n{content.strip()}"
            note = "URL → fetched title"

    if transformed is None and len(content) > 1200:
        try:
            from core.summarizer import summarize
            transformed = summarize(content, query="", mode="short", max_chars=500)
            note = "long text → summary"
        except Exception:
            pass

    if transformed is None and "\n" in content and any(
        line.startswith("    ") for line in content.split("\n")[:5]
    ):
        transformed = _trim_code(content)
        note = "code → de-indented"

    if transformed is None:
        return "Clipboard already clean — kuch reformat karne layak nahi mila."

    try:
        pyperclip.copy(transformed)
    except Exception as e:
        return f"Reformat ho gaya par clipboard pe wapas copy nahi hua: {e}"
    preview = transformed[:120].replace("\n", " ⏎ ")
    return f"Smart paste: {note}. Preview: {preview}{'…' if len(transformed) > 120 else ''}"
