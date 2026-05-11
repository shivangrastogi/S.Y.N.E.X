"""Expense tracker + life dashboard workbook.

Single source of truth: ``data/jarvis_workbook.xlsx``. Six tabs:

  Dashboard         - KPIs (month total, top category, today's spend, task count)
  Expenses          - one row per spend, with auto-derived month/week/weekday
  Tasks             - personal todo list (priority, due_date, status)
  Meetings          - scheduled meetings
  Reminders         - mirror of the live scheduler queue
  Category Summary  - pivot: category x month
  Charts            - pie + bar + line charts

Voice triggers:
  "500 rupees food pe kharch kiye"  -> add_expense
  "I spent 1200 on groceries"        -> add_expense
  "is mahine kitna kharcha"          -> month_summary
  "add task write the report by friday" -> add_task
  "schedule meeting with rohan tomorrow 5 pm" -> add_meeting
  "show expenses" / "open sheet"     -> open_workbook

Style: dark-cyan header bands, frozen panes, INR currency format,
conditional formatting (high spend -> red), alternating row stripes.
"""

from __future__ import annotations

import logging
import os
import re
import subprocess
import sys
from datetime import datetime, timedelta
from pathlib import Path
from typing import Optional

_HERE = os.path.dirname(os.path.abspath(__file__))
_ROOT = os.path.dirname(_HERE)
if _ROOT not in sys.path:
    sys.path.insert(0, _ROOT)

from core.skill_registry import skill  # noqa: E402

log = logging.getLogger(__name__)

_WORKBOOK_PATH = Path(_ROOT) / "data" / "jarvis_workbook.xlsx"

# --------------------------------------------------------------------------- #
#  Auto-categorization                                                        #
# --------------------------------------------------------------------------- #

_CATEGORY_KEYWORDS: dict[str, list[str]] = {
    "Food & Dining":   ["food", "lunch", "dinner", "breakfast", "snack",
                        "restaurant", "swiggy", "zomato", "khana", "pizza",
                        "burger", "biryani", "tea", "coffee", "chai"],
    "Groceries":       ["grocery", "groceries", "vegetables", "sabzi",
                        "dmart", "bigbasket", "blinkit", "zepto", "milk",
                        "bread", "rice", "atta", "ration"],
    "Transport":       ["uber", "ola", "rapido", "auto", "taxi", "petrol",
                        "diesel", "fuel", "metro", "train", "bus", "flight",
                        "indigo", "vistara", "cab"],
    "Utilities":       ["electricity", "bijli", "water", "internet", "wifi",
                        "phone", "mobile", "recharge", "broadband", "gas"],
    "Entertainment":   ["movie", "cinema", "netflix", "spotify", "prime",
                        "hotstar", "youtube", "concert", "game", "steam"],
    "Shopping":        ["amazon", "flipkart", "myntra", "shopping", "clothes",
                        "shoes", "electronics", "gadget"],
    "Health":          ["doctor", "hospital", "medicine", "dawai", "pharmacy",
                        "apollo", "gym", "fitness"],
    "Education":       ["course", "book", "udemy", "coursera", "tuition",
                        "school", "college", "fees"],
    "Bills & EMI":     ["emi", "loan", "credit card", "rent", "kiraya",
                        "insurance", "premium"],
    "Travel":          ["hotel", "airbnb", "trip", "travel", "vacation",
                        "ticket", "booking"],
}

_DEFAULT_CATEGORY = "Misc"


def _guess_category(text: str) -> str:
    if not text:
        return _DEFAULT_CATEGORY
    t = text.lower()
    for cat, keys in _CATEGORY_KEYWORDS.items():
        if any(k in t for k in keys):
            return cat
    return _DEFAULT_CATEGORY


# --------------------------------------------------------------------------- #
#  Slot parsing helpers                                                       #
# --------------------------------------------------------------------------- #

_AMOUNT_RE = re.compile(
    r"(?:rs\.?|inr|₹)?\s*"
    r"(\d+(?:[,]\d+)*(?:\.\d+)?)"
    r"\s*(?:rs\.?|inr|₹|rupees?|rupaye|rupay)?",
    re.IGNORECASE,
)


def _parse_amount(text: str) -> Optional[float]:
    if not text:
        return None
    m = _AMOUNT_RE.search(text)
    if not m:
        return None
    raw = m.group(1).replace(",", "")
    try:
        return float(raw)
    except ValueError:
        return None


def _parse_due_date(text: str) -> Optional[datetime]:
    """Lightweight relative-date parser. Today/tomorrow/in N days/by Friday/+ time."""
    if not text:
        return None
    t = text.lower().strip()
    today = datetime.now().replace(hour=23, minute=59, second=0, microsecond=0)

    # Resolve the DATE component first (default = today).
    base = today
    if "tomorrow" in t or "kal" in t:
        base = today + timedelta(days=1)
    elif "day after" in t or "parso" in t:
        base = today + timedelta(days=2)
    else:
        m = re.search(r"in (\d+)\s+days?", t)
        if m:
            base = today + timedelta(days=int(m.group(1)))
        else:
            weekdays = ["monday", "tuesday", "wednesday", "thursday",
                        "friday", "saturday", "sunday"]
            for i, name in enumerate(weekdays):
                if name in t:
                    delta = (i - today.weekday()) % 7
                    if delta == 0:
                        delta = 7
                    base = today + timedelta(days=delta)
                    break

    # Then merge in any "5 pm"-style time component if one exists.
    m = re.search(r"(\d{1,2})(?::(\d{2}))?\s*(am|pm)", t)
    if m:
        h = int(m.group(1))
        mn = int(m.group(2) or 0)
        if m.group(3) == "pm" and h < 12:
            h += 12
        if m.group(3) == "am" and h == 12:
            h = 0
        return base.replace(hour=h, minute=mn, second=0, microsecond=0)

    if base != today or "today" in t or "aaj" in t:
        return base
    return None


# --------------------------------------------------------------------------- #
#  Workbook builder + loader                                                  #
# --------------------------------------------------------------------------- #

# Tab schemas (column headers only — formulas are added by _refresh_).
_SHEET_HEADERS = {
    "Dashboard":        [],  # custom layout, no flat header
    "Expenses":         ["Date", "Amount (INR)", "Category", "Description",
                         "Payment Method", "Month", "Week", "Weekday"],
    "Tasks":            ["Created", "Task", "Priority", "Due Date",
                         "Status", "Notes"],
    "Meetings":         ["Created", "Person", "Meeting Time", "Agenda",
                         "Location", "Status"],
    "Reminders":        ["Set At", "Fire At", "Message", "Status"],
    "Category Summary": ["Category", "This Month (INR)", "Last Month (INR)",
                         "Trend", "% of Total"],
}


_HEADER_FILL = "FF064E5C"     # deep teal
_HEADER_FONT_HEX = "FFEAF6F8" # near-white
_STRIPE_FILL = "FFF1F8FA"     # subtle alt row
_TODAY_FILL = "FFFFF1B0"      # gentle yellow


def _try_import_openpyxl():
    try:
        import openpyxl  # noqa: F401
        from openpyxl import Workbook, load_workbook  # noqa: F401
        from openpyxl.chart import BarChart, LineChart, PieChart, Reference  # noqa: F401
        from openpyxl.styles import (Alignment, Border, Font, PatternFill,
                                     Side)  # noqa: F401
        from openpyxl.utils import get_column_letter  # noqa: F401
        return True
    except Exception:
        return False


def _ensure_workbook():
    """Open existing or create a fresh styled workbook. Returns the openpyxl wb."""
    if not _try_import_openpyxl():
        return None

    from openpyxl import Workbook, load_workbook

    if _WORKBOOK_PATH.exists():
        try:
            return load_workbook(_WORKBOOK_PATH)
        except Exception as e:
            log.warning("[expense_tracker] workbook corrupt (%s); rebuilding", e)

    _WORKBOOK_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    # Drop the default sheet — we recreate everything in order.
    default = wb.active
    wb.remove(default)
    for name, headers in _SHEET_HEADERS.items():
        ws = wb.create_sheet(name)
        if headers:
            ws.append(headers)
    _style_workbook(wb)
    _build_dashboard(wb)
    wb.save(_WORKBOOK_PATH)
    return wb


def _style_workbook(wb) -> None:
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    header_fill = PatternFill("solid", fgColor=_HEADER_FILL)
    header_font = Font(bold=True, color=_HEADER_FONT_HEX, size=11)
    align = Alignment(horizontal="center", vertical="center")

    column_widths = {
        "Expenses":         [14, 16, 22, 38, 18, 12, 8, 12],
        "Tasks":            [20, 50, 12, 18, 14, 30],
        "Meetings":         [20, 24, 22, 40, 22, 14],
        "Reminders":        [20, 22, 50, 14],
        "Category Summary": [22, 18, 18, 14, 14],
    }

    for name, widths in column_widths.items():
        ws = wb[name]
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w
        # Style the header row.
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = align
        ws.row_dimensions[1].height = 24
        ws.freeze_panes = "A2"


def _build_dashboard(wb) -> None:
    """Lay out the Dashboard tab with KPI tiles + headers."""
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    ws = wb["Dashboard"]
    ws.column_dimensions["A"].width = 4
    for col in "BCDEFG":
        ws.column_dimensions[col].width = 22

    title_fill = PatternFill("solid", fgColor=_HEADER_FILL)
    tile_fill = PatternFill("solid", fgColor="FFE6F4F1")
    title_font = Font(bold=True, color=_HEADER_FONT_HEX, size=18)
    tile_label_font = Font(bold=True, color="FF064E5C", size=10)
    tile_value_font = Font(bold=True, color="FF0F172A", size=22)
    side = Side(style="thin", color="FFB7DDE2")
    border = Border(left=side, right=side, top=side, bottom=side)

    ws.merge_cells("B2:G2")
    ws["B2"] = "  AERIS — Life & Expense Dashboard"
    ws["B2"].fill = title_fill
    ws["B2"].font = title_font
    ws["B2"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 36

    # KPI tiles (label + formula)
    tiles = [
        ("B4", "B5", "This Month Spend",
            "=IFERROR(SUMPRODUCT((Expenses!F2:F1000=TEXT(TODAY(),\"mmm yy\"))*Expenses!B2:B1000),0)"),
        ("D4", "D5", "Today Spend",
            "=IFERROR(SUMPRODUCT((Expenses!A2:A1000=TODAY())*Expenses!B2:B1000),0)"),
        ("F4", "F5", "Open Tasks",
            "=COUNTIF(Tasks!E2:E1000,\"open\")"),
        ("B7", "B8", "Total Logged",
            "=COUNTA(Expenses!A2:A1000)"),
        ("D7", "D8", "Top Category (Month)",
            "=IFERROR(INDEX('Category Summary'!A2:A50,"
            "MATCH(MAX('Category Summary'!B2:B50),'Category Summary'!B2:B50,0)),\"-\")"),
        ("F7", "F8", "Upcoming Meetings",
            "=COUNTIF(Meetings!F2:F1000,\"scheduled\")"),
    ]
    for label_cell, val_cell, label, formula in tiles:
        ws[label_cell] = label
        ws[label_cell].font = tile_label_font
        ws[label_cell].fill = tile_fill
        ws[label_cell].alignment = Alignment(horizontal="center", vertical="center")
        ws[label_cell].border = border
        ws[val_cell] = formula
        ws[val_cell].font = tile_value_font
        ws[val_cell].fill = tile_fill
        ws[val_cell].alignment = Alignment(horizontal="center", vertical="center")
        ws[val_cell].border = border
        # Currency formatting only for the spend tiles.
        if "Spend" in label or "Logged" in label and label != "Total Logged":
            ws[val_cell].number_format = '"₹"#,##0'
        elif "Spend" in label:
            ws[val_cell].number_format = '"₹"#,##0'
        ws.row_dimensions[int(label_cell[1:])].height = 22
        ws.row_dimensions[int(val_cell[1:])].height = 36

    ws.merge_cells("B10:G10")
    ws["B10"] = "  Charts auto-refresh from the Expenses tab — see 'Charts' sheet."
    ws["B10"].font = Font(italic=True, color="FF475569")

    ws.sheet_view.showGridLines = False


def _refresh_styles_and_charts(wb) -> None:
    """Re-apply alt-row stripes, currency format, and rebuild charts."""
    from openpyxl.chart import BarChart, LineChart, PieChart, Reference
    from openpyxl.chart.label import DataLabelList
    from openpyxl.formatting.rule import ColorScaleRule
    from openpyxl.styles import PatternFill

    # Stripe + currency on Expenses
    ws_exp = wb["Expenses"]
    stripe = PatternFill("solid", fgColor=_STRIPE_FILL)
    today_fill = PatternFill("solid", fgColor=_TODAY_FILL)
    today_str = datetime.now().date().isoformat()

    last_row = max(ws_exp.max_row, 1)
    for r in range(2, last_row + 1):
        if r % 2 == 0:
            for c in range(1, len(_SHEET_HEADERS["Expenses"]) + 1):
                ws_exp.cell(row=r, column=c).fill = stripe
        # Today highlight
        date_cell = ws_exp.cell(row=r, column=1)
        if date_cell.value and str(date_cell.value).startswith(today_str):
            for c in range(1, len(_SHEET_HEADERS["Expenses"]) + 1):
                ws_exp.cell(row=r, column=c).fill = today_fill
        # Currency on amount column
        amt = ws_exp.cell(row=r, column=2)
        if amt.value not in (None, ""):
            amt.number_format = '"₹"#,##0.00'

    # Conditional gradient on Amount (col B) so big spends pop red.
    if last_row >= 2:
        rng = f"B2:B{last_row}"
        ws_exp.conditional_formatting.add(
            rng,
            ColorScaleRule(
                start_type="min", start_color="FF63B3A8",
                mid_type="percentile", mid_value=50, mid_color="FFFFD972",
                end_type="max", end_color="FFE15A5A",
            ),
        )

    # Rebuild Charts tab from scratch each refresh.
    from openpyxl.styles import Font as _Font
    if "Charts" in wb.sheetnames:
        wb.remove(wb["Charts"])
    ws_chart = wb.create_sheet("Charts")
    ws_chart.sheet_view.showGridLines = False
    ws_chart["B2"] = "Spend Insights"
    # Fresh Font object — assigning a StyleProxy from another cell raises
    # "unhashable type: StyleProxy" at save time.
    ws_chart["B2"].font = _Font(bold=True, color=_HEADER_FONT_HEX, size=18)
    ws_chart["B2"].fill = PatternFill("solid", fgColor=_HEADER_FILL)

    # Recompute Category Summary first.
    _refresh_category_summary(wb)

    cs_ws = wb["Category Summary"]
    cat_last = cs_ws.max_row
    if cat_last >= 2:
        # Pie: this-month spend by category
        pie = PieChart()
        labels = Reference(cs_ws, min_col=1, min_row=2, max_row=cat_last)
        data = Reference(cs_ws, min_col=2, min_row=1, max_row=cat_last)
        pie.add_data(data, titles_from_data=True)
        pie.set_categories(labels)
        pie.title = "This Month — Category Mix"
        pie.height = 9
        pie.width = 14
        pie.dataLabels = DataLabelList(showPercent=True)
        ws_chart.add_chart(pie, "B4")

        # Bar: this-month vs last-month per category
        bar = BarChart()
        bar.type = "col"
        bar.style = 11
        bar.title = "This Month vs Last Month"
        bar.y_axis.title = "INR"
        bar.x_axis.title = "Category"
        bar_data = Reference(cs_ws, min_col=2, max_col=3,
                             min_row=1, max_row=cat_last)
        bar_cats = Reference(cs_ws, min_col=1, min_row=2, max_row=cat_last)
        bar.add_data(bar_data, titles_from_data=True)
        bar.set_categories(bar_cats)
        bar.height = 9
        bar.width = 18
        ws_chart.add_chart(bar, "L4")

    # Line chart: cumulative spend across this month's expenses.
    if last_row >= 2:
        line = LineChart()
        line.title = "Daily Spend (Expenses tab)"
        line.y_axis.title = "INR"
        line.x_axis.title = "Date"
        data = Reference(ws_exp, min_col=2, min_row=1, max_row=last_row)
        cats = Reference(ws_exp, min_col=1, min_row=2, max_row=last_row)
        line.add_data(data, titles_from_data=True)
        line.set_categories(cats)
        line.height = 9
        line.width = 18
        ws_chart.add_chart(line, "B24")


def _refresh_category_summary(wb) -> None:
    """Rebuild the Category Summary tab from raw Expenses rows."""
    ws_exp = wb["Expenses"]
    cs = wb["Category Summary"]

    # Wipe data rows but keep the header.
    if cs.max_row > 1:
        cs.delete_rows(2, cs.max_row)

    if ws_exp.max_row < 2:
        return

    now = datetime.now()
    this_label = now.strftime("%b %y")
    last_month_dt = (now.replace(day=1) - timedelta(days=1))
    last_label = last_month_dt.strftime("%b %y")

    by_cat_this: dict[str, float] = {}
    by_cat_last: dict[str, float] = {}
    grand_this = 0.0

    for row in ws_exp.iter_rows(min_row=2, values_only=True):
        if not row or row[0] in (None, ""):
            continue
        amt = row[1] if isinstance(row[1], (int, float)) else _parse_amount(str(row[1]))
        if amt is None:
            continue
        cat = (row[2] or _DEFAULT_CATEGORY).strip()
        month_label = row[5] or ""
        if month_label == this_label:
            by_cat_this[cat] = by_cat_this.get(cat, 0) + amt
            grand_this += amt
        elif month_label == last_label:
            by_cat_last[cat] = by_cat_last.get(cat, 0) + amt

    all_cats = sorted(set(by_cat_this) | set(by_cat_last))
    for cat in all_cats:
        this_v = by_cat_this.get(cat, 0)
        last_v = by_cat_last.get(cat, 0)
        if last_v > 0:
            trend = "+" if this_v >= last_v else "-"
            trend = f"{trend}{abs(this_v - last_v) / last_v * 100:.0f}%"
        elif this_v > 0:
            trend = "NEW"
        else:
            trend = "-"
        share = f"{this_v / grand_this * 100:.0f}%" if grand_this > 0 else "0%"
        cs.append([cat, round(this_v, 2), round(last_v, 2), trend, share])

    # Currency format on numeric columns.
    for r in range(2, cs.max_row + 1):
        for col in (2, 3):
            cs.cell(row=r, column=col).number_format = '"₹"#,##0'


# --------------------------------------------------------------------------- #
#  Public skill handlers                                                      #
# --------------------------------------------------------------------------- #

def _ensure_or_message():
    if not _try_import_openpyxl():
        return None, ("openpyxl install nahi hai, sir. "
                      "'pip install openpyxl' chalao to expense tracking on ho jayegi.")
    wb = _ensure_workbook()
    return wb, None


def _safe_save(wb) -> Optional[str]:
    """Save with one retry. Returns None on success, an error string on failure.

    PermissionError almost always means the user has the file open in Excel —
    we surface a friendly message instead of letting the exception bubble.
    """
    try:
        wb.save(_WORKBOOK_PATH)
        return None
    except PermissionError:
        return ("Workbook abhi Excel mein open hai, sir — close karke phir try karo. "
                "Data safe hai, bas save block ho gaya.")
    except Exception as e:
        return f"Save mein dikkat: {e}"


@skill(
    name="add_expense",
    description="Log a new expense (amount + description) into the local workbook with auto category.",
    patterns=[
        # English-style
        "spent 500 on food", "i spent 1200 on groceries",
        "log expense 1500 grocery", "add expense 700 dinner",
        "i paid 800 for movie", "paid 250 for chai",
        "track expense 600 medicine",
        # Hindi/Hinglish "kharch" verb
        "500 rupees food pe kharch kiye", "200 rs uber pe lagaye",
        "1000 rupaye shopping mein kharch hua", "300 rupaye chai pe lagaye",
        "kharcha note karo 800 movie", "kharcha 1500 amazon pe",
        # "ka X lagaya / ka X kiya / ka X liya" — common Hinglish expense verbs
        "750 ka uber lagaya", "500 ka swiggy order kiya",
        "1200 ka grocery liya", "300 ka petrol bharaya",
        "400 ka fuel diya", "200 ka chai kiya",
        "1500 ka amazon order kiya", "800 ka dinner kiya",
        "250 ka auto liya", "600 ka medicine liya",
        # Imperative
        "expense add karo 250 chai", "expense log karo 800",
        "track kar lo 400 fuel", "kharch add kar 500 food",
        "add karo 1000 shopping", "log kar do 350 dinner",
        # Generic spend
        "maine 500 rupaye uber pe diye", "main ne 800 lagaye netflix",
        "1000 ke shopping kiye", "200 ki chai pi",
    ],
    required_entities=["amount", "description"],
    prompts={
        "amount": "Kitne rupaye ka kharcha tha?",
        "description": "Kis cheez pe kharcha hua?",
    },
)
def add_expense(slots: dict) -> str:
    wb, err = _ensure_or_message()
    if err:
        return err

    raw = " ".join(str(v) for v in slots.values() if v)
    amount = _parse_amount(slots.get("amount") or raw)
    desc = (slots.get("description") or "").strip()
    if not desc:
        # Heuristic: drop the amount tokens and keep the rest as description.
        desc = re.sub(_AMOUNT_RE, " ", raw).strip()
        desc = re.sub(r"\b(rupees?|rupaye|rs|inr|on|pe|par|ka|ki|ke|spent|kharch[a-z]*|"
                      r"i|main|maine|expense|log|add|note|karo)\b", " ", desc, flags=re.I)
        desc = re.sub(r"\s+", " ", desc).strip(" .,-")
    if amount is None:
        return "Amount nahi pakad paya — number repeat kar do, sir. (jaise '500 rupees food pe')"
    if not desc:
        desc = "(unspecified)"

    category = _guess_category(desc)
    now = datetime.now()
    ws = wb["Expenses"]
    ws.append([
        now.date(),
        round(amount, 2),
        category,
        desc,
        "Cash",                       # default; user can edit in sheet
        now.strftime("%b %y"),
        f"W{now.isocalendar()[1]:02d}",
        now.strftime("%a"),
    ])
    try:
        _refresh_styles_and_charts(wb)
    except Exception as e:
        log.info("[expense_tracker] chart refresh failed (non-fatal): %s", e)
    err = _safe_save(wb)
    if err:
        return err
    return (f"Logged: ₹{amount:.0f} on {desc} ({category}). "
            f"Sheet update ho gayi, sir.")


@skill(
    name="month_summary",
    description="Speak the current month's spending breakdown.",
    patterns=[
        "is mahine kitna kharcha", "month ka kharcha kitna hai",
        "this month spending", "monthly summary do",
        "kharcha summary batao", "expense summary",
        "is mahine kya kharch kiya", "show monthly expenses",
    ],
    required_entities=[],
)
def month_summary(slots: dict) -> str:
    wb, err = _ensure_or_message()
    if err:
        return err

    _refresh_category_summary(wb)
    _safe_save(wb)  # best-effort; if Excel has it open we still read in-memory

    cs = wb["Category Summary"]
    if cs.max_row < 2:
        return "Is mahine abhi tak koi expense nahi log hua, sir."

    rows = list(cs.iter_rows(min_row=2, values_only=True))
    total = sum(r[1] for r in rows if isinstance(r[1], (int, float)))
    rows.sort(key=lambda r: r[1] if isinstance(r[1], (int, float)) else 0, reverse=True)
    top = rows[0]
    top_cat, top_val = top[0], top[1]

    breakdown = ", ".join(
        f"{r[0]} ₹{int(r[1])}" for r in rows[:4]
        if isinstance(r[1], (int, float)) and r[1] > 0
    )
    return (f"Is mahine total ₹{int(total)} kharch hua. "
            f"Top category: {top_cat} (₹{int(top_val)}). "
            f"Breakdown: {breakdown}.")


@skill(
    name="add_task",
    description="Add a personal task to the Tasks tab with priority + due date.",
    patterns=[
        "add task write the report", "task add karo project finish karna hai",
        "remind me to call mom tomorrow", "todo add karo gym jana",
        "kaam add karo", "important kaam add karo",
        "task list mein add karo", "schedule task",
        "add to my todo list", "note this task",
    ],
    required_entities=["task_text"],
    prompts={"task_text": "Kya task add karna hai?"},
)
def add_task(slots: dict) -> str:
    wb, err = _ensure_or_message()
    if err:
        return err

    task_text = (slots.get("task_text") or "").strip()
    if not task_text:
        return "Task content nahi mila — kya likhna hai batao."

    due = _parse_due_date(task_text)
    priority = "High" if any(k in task_text.lower()
                             for k in ("urgent", "asap", "important", "zaroori")) else "Medium"
    now = datetime.now()
    ws = wb["Tasks"]
    ws.append([
        now.strftime("%Y-%m-%d %H:%M"),
        task_text,
        priority,
        due.strftime("%Y-%m-%d %H:%M") if due else "",
        "Open",
        "",
    ])
    err = _safe_save(wb)
    if err:
        return err
    due_str = due.strftime("%a, %d %b") if due else "no deadline"
    return f"Task add kar diya: '{task_text}' (priority {priority}, due {due_str})."


@skill(
    name="add_meeting",
    description="Add a meeting entry to the Meetings tab.",
    patterns=[
        "schedule meeting with rohan tomorrow 5 pm",
        "meeting add karo with team monday",
        "rohan ke saath meeting tomorrow 10 am",
        "add meeting", "calendar mein meeting add karo",
        "log a meeting", "meeting note karo",
    ],
    required_entities=["person", "time"],
    prompts={
        "person": "Kis ke saath meeting hai?",
        "time": "Meeting kab hai?",
    },
)
def add_meeting(slots: dict) -> str:
    wb, err = _ensure_or_message()
    if err:
        return err

    person = (slots.get("person") or "Unknown").strip()
    time_str = (slots.get("time") or "").strip()
    agenda = (slots.get("agenda") or slots.get("description") or "").strip()
    when = _parse_due_date(time_str) if time_str else None

    now = datetime.now()
    ws = wb["Meetings"]
    ws.append([
        now.strftime("%Y-%m-%d %H:%M"),
        person,
        when.strftime("%Y-%m-%d %H:%M") if when else (time_str or "TBD"),
        agenda or "",
        "",
        "Scheduled",
    ])
    err = _safe_save(wb)
    if err:
        return err
    return (f"Meeting scheduled: {person} — "
            f"{when.strftime('%a, %d %b %I:%M %p') if when else (time_str or 'TBD')}.")


@skill(
    name="open_workbook",
    description="Open the local jarvis workbook (expenses, tasks, meetings) in Excel.",
    patterns=[
        "open expense sheet", "open jarvis workbook",
        "show expenses", "expense sheet kholo",
        "show my sheet", "open my dashboard sheet",
        "kharcha sheet kholo", "excel kholo expenses ka",
        "workbook kholo", "show workbook",
    ],
    required_entities=[],
)
def open_workbook(slots: dict) -> str:
    wb, err = _ensure_or_message()
    if err:
        return err
    try:
        wb.save(_WORKBOOK_PATH)
    except Exception:
        pass
    try:
        os.startfile(str(_WORKBOOK_PATH))
    except AttributeError:
        # Non-Windows fallback (shouldn't happen on this project but harmless).
        subprocess.Popen(["xdg-open", str(_WORKBOOK_PATH)])
    except Exception as e:
        return f"Sheet kholne mein dikkat: {e}. Path: {_WORKBOOK_PATH}"
    return f"Workbook khol raha hoon: {_WORKBOOK_PATH.name}"


# --------------------------------------------------------------------------- #
#  Smoke test                                                                 #
# --------------------------------------------------------------------------- #

if __name__ == "__main__":
    logging.basicConfig(level=logging.INFO, format="%(message)s")
    print(add_expense({"amount": "500 rupees", "description": "swiggy biryani"}))
    print(add_expense({"amount": "1200", "description": "uber to office"}))
    print(add_task({"task_text": "finish brain refactor by friday urgent"}))
    print(add_meeting({"person": "rohan", "time": "tomorrow 5 pm",
                       "agenda": "review jarvis"}))
    print(month_summary({}))
